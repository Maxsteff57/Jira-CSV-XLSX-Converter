"""
Конвертер Jira CSV ↔ XLSX
─────────────────────────
Десктопное приложение для преобразования экспорта задач Jira (CSV) в
форматированные Excel-таблицы. Поддерживает:
  • drag & drop
  • Очистка названий колонок
  • фильтрацию и переупорядочивание через columns.txt
  • стили таблиц Excel (превью + результат)
  • светлую/тёмную тему с автоопределением из реестра Windows
  • локализацию на 8 языков

Автор: maxsteff  ·  https://t.me/maxsteff
"""
import sys, re, os, json, threading, winreg
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from pathlib import Path

# tkinterdnd2 опционален — если нет, drag & drop просто отключится
try:
    from tkinterdnd2 import TkinterDnD, DND_FILES
    _DND = True
except ImportError:
    _DND = False

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ─── Константы ────────────────────────────────────────────────────────────────

FONT          = "Segoe UI"
MIN_W         = 600                              # минимальная ширина окна
APP_NAME      = "Конвертер Jira CSV ↔ XLSX"
VERSION       = "0.1 beta"
SETTINGS_FILE = "settings.json"                  # хранится рядом с .exe

# Универсальная карта сокращений месяцев (8 языков) → номер месяца.
# Ключи приведены к нижнему регистру, берем первые 3 символа.
MONTH_MAP = {
    # Русский (ru)
    "янв": "01", "фев": "02", "мар": "03", "апр": "04", "май": "05", "мая": "05",
    "июн": "06", "июл": "07", "авг": "08", "сен": "09", "окт": "10", "ноя": "11", "дек": "12",
    # English (en) / Deutsch (de) / Français (fr) / Português (pt)
    "jan": "01", "feb": "02", "mar": "03", "apr": "04", "may": "05", "mai": "05",
    "jun": "06", "jul": "07", "aug": "08", "sep": "09", "oct": "10", "nov": "11", "dec": "12",
    # Дополнительно для Deutsch/Français/Português/Español
    "mär": "03", "avr": "04", "jui": "06", "aoû": "08", "dez": "12", "fev": "02", 
    "set": "09", "out": "10", "déc": "12", "fév": "02",
    # Español (es)
    "ene": "01", "abr": "04", "ago": "08", "dic": "12"
    # Примечание: zh (中文) и ar (العربية) в CSV Jira часто используют 
    # либо числовой формат (01, 02...), либо стандартные латинские сокращения (Jan, Feb...).
}

# ─── Стили таблицы Excel ─────────────────────────────────────────────────────
# Структура кортежа: (key, xlsx_name, group, hdr, odd, even, hdr_text)
#   key       — внутренний идентификатор, не зависит от языка (хранится в settings)
#   xlsx_name — имя встроенного стиля openpyxl/Excel (None = без стиля)
#   group     — секция в попапе ("none"/"light"/"medium"/"dark")
#   hdr       — заливка заголовка
#   odd/even  — заливка нечётных/чётных строк (для миниатюры в попапе)
#   hdr_text  — цвет текста заголовка
#
# ВАЖНО про цвета: openpyxl использует тему Office 2007 по умолчанию, поэтому
# Excel рендерит таблицы в "старой" палитре (синий 4F81BD, зелёный 9BBB59, …),
# а не в современной Office 2016 (синий 4472C4, зелёный 70AD47).
# Цвета миниатюр должны соответствовать реальному результату — поэтому здесь
# именно Office 2007.
ALL_TABLE_STYLES = [
    # key             xlsx_name             group     hdr        odd        even       hdr_txt
    ("s_none",        None,                 "none",   "#F2F2F2", "#FFFFFF", "#FFFFFF", "#000000"),

    # ── Светлые ── тонкая цветная полоса вверху, белый/lighter в строках
    ("s_l_black",     "TableStyleLight8",   "light",  "#000000", "#FFFFFF", "#D9D9D9", "#FFFFFF"),
    ("s_l_blue",      "TableStyleLight9",   "light",  "#4F81BD", "#FFFFFF", "#DCE6F1", "#FFFFFF"),
    ("s_l_red",       "TableStyleLight10",  "light",  "#C0504D", "#FFFFFF", "#F2DCDB", "#FFFFFF"),
    ("s_l_green",     "TableStyleLight11",  "light",  "#9BBB59", "#FFFFFF", "#EBF1DE", "#FFFFFF"),
    ("s_l_purple",    "TableStyleLight12",  "light",  "#8064A2", "#FFFFFF", "#E4DFEC", "#FFFFFF"),
    ("s_l_aqua",      "TableStyleLight13",  "light",  "#4BACC6", "#FFFFFF", "#DAEEF3", "#FFFFFF"),
    ("s_l_orange",    "TableStyleLight14",  "light",  "#F79646", "#FFFFFF", "#FDE9D9", "#FFFFFF"),

    # ── Средние ── насыщенный заголовок + чередование mid/light
    ("s_m_black",     "TableStyleMedium1",  "medium", "#000000", "#A6A6A6", "#D9D9D9", "#FFFFFF"),
    ("s_m_blue",      "TableStyleMedium2",  "medium", "#4F81BD", "#B8CCE4", "#DCE6F1", "#FFFFFF"),
    ("s_m_red",       "TableStyleMedium3",  "medium", "#C0504D", "#E5B9B7", "#F2DCDB", "#FFFFFF"),
    ("s_m_green",     "TableStyleMedium4",  "medium", "#9BBB59", "#D7E3BC", "#EBF1DE", "#FFFFFF"),
    ("s_m_purple",    "TableStyleMedium5",  "medium", "#8064A2", "#CCC1D9", "#E4DFEC", "#FFFFFF"),
    ("s_m_aqua",      "TableStyleMedium6",  "medium", "#4BACC6", "#B7DDE8", "#DAEEF3", "#FFFFFF"),
    ("s_m_orange",    "TableStyleMedium7",  "medium", "#F79646", "#FCD5B4", "#FDE9D9", "#FFFFFF"),

    # ── Тёмные ── глубокий заголовок + accent в полосах
    ("s_d_black",     "TableStyleDark1",    "dark",   "#000000", "#595959", "#7F7F7F", "#FFFFFF"),
    ("s_d_blue",      "TableStyleDark2",    "dark",   "#17365D", "#4F81BD", "#366092", "#FFFFFF"),
    ("s_d_red",       "TableStyleDark3",    "dark",   "#632423", "#C0504D", "#953734", "#FFFFFF"),
    ("s_d_green",     "TableStyleDark4",    "dark",   "#4F6128", "#9BBB59", "#76923C", "#FFFFFF"),
    ("s_d_purple",    "TableStyleDark5",    "dark",   "#3F3151", "#8064A2", "#5F497A", "#FFFFFF"),
    ("s_d_aqua",      "TableStyleDark6",    "dark",   "#205867", "#4BACC6", "#31859B", "#FFFFFF"),
    ("s_d_orange",    "TableStyleDark7",    "dark",   "#E26B0A", "#F79646", "#974806", "#FFFFFF"),
]

# Производные словари для быстрого доступа.
# *_BY_KEY:  key → openpyxl имя (то что записывается в xlsx)
# *_COLORS:  key → (hdr, odd, even, hdr_text) — цвета для отрисовки превью
TABLE_STYLE_BY_KEY = {s[0]: s[1] for s in ALL_TABLE_STYLES}
TABLE_STYLE_COLORS = {s[0]: s[3:] for s in ALL_TABLE_STYLES}

def get_table_style_names(lang: str) -> list:
    """Список локализованных названий стилей в порядке ALL_TABLE_STYLES."""
    d = STRINGS.get(lang, STRINGS["en"])
    return [d.get(s[0], s[0]) for s in ALL_TABLE_STYLES]


def get_style_key_by_label(label: str, lang: str) -> str | None:
    """Обратное преобразование: «Средний — синий» → 's_m_blue'."""
    d = STRINGS.get(lang, STRINGS["en"])
    for s in ALL_TABLE_STYLES:
        if d.get(s[0], s[0]) == label:
            return s[0]
    return None


# ─── Темы оформления ──────────────────────────────────────────────────────────
# Палитра цветов для светлого и тёмного режимов.
# Расшифровка: BG — фон окна, CARD — карточка с полями, ACCENT — шапка,
# ACC2 — акцентные кнопки/прогресс, ENTRY — поля ввода, DZ_* — drop-zone,
# PB_BG — фон прогресс-бара.
THEMES = {
    "light": {
        "BG":     "#F0F4FA",
        "CARD":   "#FFFFFF",
        "ACCENT": "#1F3864",
        "ACC2":   "#2E5FA3",
        "TEXT":   "#1A1A2E",
        "MUTED":  "#7F8C8D",
        "BORDER": "#D5DCE8",
        "ENTRY":  "#F7F9FC",
        "DZ_BG":  "#EEF3FB",
        "DZ_CL":  "#A8C0E8",
        "DZ_TXT": "#5A8AD0",
        "PB_BG":  "#DDE4EE",
    },
    "dark": {
        "BG":     "#1E1E2E",
        "CARD":   "#2A2A3E",
        "ACCENT": "#0F1F40",
        "ACC2":   "#3A6FBF",
        "TEXT":   "#E0E0F0",
        "MUTED":  "#8A8AAA",
        "BORDER": "#3A3A5A",
        "ENTRY":  "#252535",
        "DZ_BG":  "#252540",
        "DZ_CL":  "#4A6090",
        "DZ_TXT": "#6080B0",
        "PB_BG":  "#2A2A45",
    },
}

# Универсальные акценты (одинаковы в обеих темах) — статус успеха/ошибки.
GREEN = "#1E8449"
RED   = "#C0392B"


# ─── Локализация ──────────────────────────────────────────────────────────────
# LANGUAGES — что показывается в селекторе языка.
# STRINGS   — словарь {lang_code: {key: translation}}, см. функцию s() ниже.

LANGUAGES = {
    "ru": "Русский",
    "en": "English",
    "de": "Deutsch",
    "fr": "Français",
    "es": "Español",
    "zh": "中文",
    "ar": "العربية",
    "pt": "Português",
}

STRINGS = {
    "ru": {
        "title":          "Конвертер Jira CSV ↔ XLSX",
        "subtitle":       "Jira tasks to excel to Jira tasks ",
        "about_btn":      "ℹ О программе",
        "lang_btn":       "🌐 Язык",
        "csv_lbl":        "CSV-файл (экспорт Jira):",
        "xlsx_lbl":       "Сохранить XLSX как:",
        "col_lbl":        "Фильтр колонок (columns.txt):",
        "style_lbl":      "Стиль таблицы Excel:",
        "browse":         "Обзор…",
        "create_ex":      "Создать пример",
        "convert_btn":    "  ▶  Преобразовать",
        "status_idle":    "Выберите файлы и нажмите «Преобразовать»",
        "status_working": "Преобразование…",
        "hint_none":      "Файл не выбран — будут сохранены все колонки",
        "hint_found":     "✓ Найден: ",
        "hint_cols":      " колонок",
        "hint_strips":    " префиксов обрезки",
        "hint_all":       "все колонки",
        "dz_idle":        "Перетащите сюда .csv файл  (или кликните)",
        "dz_drop":        "Отпустите файл здесь",
        "no_file":        "Выберите CSV-файл.",
        "no_path":        "Укажите путь для сохранения XLSX.",
        "no_csv_ex":      "Сначала выберите CSV-файл.",
        "file_not_found": "Файл не найден:",
        "wrong_fmt":      "Пожалуйста, выберите файл .csv",
        "wrong_fmt_ttl":  "Неверный формат",
        "done_ttl":       "Готово!",
        "done_msg":       "Файл сохранён:\n{path}\n\nОткрыть папку?",
        "done_status":    "✓ Готово: {rows} строк, {cols} колонок → {name}",
        "err_ttl":        "Ошибка преобразования",
        "err_status":     "Ошибка: {error}",
        "already_exists": "Файл уже существует:\n{path}",
        "exists_ttl":     "Уже существует",
        "created_ttl":    "Готово",
        "created_msg":    "Файл создан:\n{path}\n\nОткрыть папку?",
        "create_err":     "Не удалось создать файл.",
        "create_err_ttl": "Ошибка",
        "about_ttl":      "О программе",
        "about_ver":      "Версия {ver}  |  maxsteff",
        "about_desc": (
            "Программа предназначена для преобразования файлов\n"
            "экспорта задач из Jira (формат CSV) в Excel (.xlsx).\n\n"
            "Возможности:\n"
            "  • Перетаскивание CSV-файлов в окно программы\n"
            "  • Настраиваемая очистка названий колонок\n"
            "    (убирает \'Пользовательское поле (…)\')\n"
            "  • Фильтрация колонок через файл columns.txt\n"
            "  • Настраиваемые префиксы обрезки [strip]\n"
            "  • Конвертация дат Jira → ДД.ММ.ГГГГ\n"
            "  • Выбор стиля таблицы Excel\n"
            "  • Загрузка нескольких .csv для объединения в .xlsx\n\n"
            "Формат columns.txt:\n"
            "  [columns] — список нужных колонок\n"
            "  [strip]   — префиксы для обрезки\n"
            "  # строки с # — комментарии"
        ),
        "about_tg":       "Связь с автором: t.me/maxsteff",
        "close":          "Закрыть",
        "lang_ttl":       "Выбор языка",
        "lang_auto":      "(определён автоматически)",
        "apply":          "Применить",

        "s_none": "Нет стиля",
        "s_l_black": "Светлый — чёрный",
        "s_l_blue": "Светлый — синий",
        "s_l_red": "Светлый — красный",
        "s_l_green": "Светлый — зелёный",
        "s_l_purple": "Светлый — фиолетовый",
        "s_l_aqua": "Светлый — бирюзовый",
        "s_l_orange": "Светлый — оранжевый",
        "s_m_black": "Средний — чёрный",
        "s_m_blue": "Средний — синий",
        "s_m_red": "Средний — красный",
        "s_m_green": "Средний — зелёный",
        "s_m_purple": "Средний — фиолетовый",
        "s_m_aqua": "Средний — бирюзовый",
        "s_m_orange": "Средний — оранжевый",
        "s_d_black": "Тёмный — чёрный",
        "s_d_blue": "Тёмный — синий",
        "s_d_red": "Тёмный — красный",
        "s_d_green": "Тёмный — зелёный",
        "s_d_purple": "Тёмный — фиолетовый",
        "s_d_aqua": "Тёмный — бирюзовый",
        "s_d_orange": "Тёмный — оранжевый",
        "pick_csv_ttl":   "Выберите CSV-файл",
        "pick_xlsx_ttl":  "Сохранить XLSX как",
        "pick_col_ttl":   "Выберите columns.txt",
        "csv_ft":         "CSV файлы",
        "xlsx_ft":        "Excel файлы",
        "txt_ft":         "Текстовые файлы",
        "all_ft":         "Все файлы",
        "xlsx_in_lbl":    "XLSX-файл (после редактирования):",
        "csv_out_lbl":    "Сохранить CSV как (для Jira):",
        "rh_only_lbl":    "Высота строк:",
        "dlm_only_lbl":   "Разделитель CSV:",
        "dlm_ttl":        "Разделитель CSV",
        "dlm_msg":        "Выберите разделитель для итогового файла:",
        "dlm_ok":         "Продолжить",
        "dlm_semi":       "Точка с запятой ( ; )",
        "dlm_comma":      "Запятая ( , )",
        "dlm_pipe":       "Вертикальная черта ( | )",
        "dlm_tab":        "Знак табуляции ( Tab )",
    },
    "en": {
        "title":          "Jira CSV ↔ XLSX Converter",
        "subtitle":       "Jira tasks to excel to Jira tasks ",
        "about_btn":      "ℹ About",
        "lang_btn":       "🌐 Language",
        "csv_lbl":        "CSV file (Jira export):",
        "xlsx_lbl":       "Save XLSX as:",
        "col_lbl":        "Column filter (columns.txt):",
        "style_lbl":      "Excel table style:",
        "browse":         "Browse…",
        "create_ex":      "Create example",
        "convert_btn":    "  ▶  Convert",
        "status_idle":    "Select files and click Convert",
        "status_working": "Converting…",
        "hint_none":      "No file selected — all columns will be saved",
        "hint_found":     "✓ Found: ",
        "hint_cols":      " columns",
        "hint_strips":    " strip prefixes",
        "hint_all":       "all columns",
        "dz_idle":        "Drop a .csv file here  (or click)",
        "dz_drop":        "Release file here",
        "no_file":        "Please select a CSV file.",
        "no_path":        "Please specify where to save the XLSX.",
        "no_csv_ex":      "Please select a CSV file first.",
        "file_not_found": "File not found:",
        "wrong_fmt":      "Please select a .csv file",
        "wrong_fmt_ttl":  "Wrong format",
        "done_ttl":       "Done!",
        "done_msg":       "File saved:\n{path}\n\nOpen folder?",
        "done_status":    "✓ Done: {rows} rows, {cols} columns → {name}",
        "err_ttl":        "Conversion error",
        "err_status":     "Error: {error}",
        "already_exists": "File already exists:\n{path}",
        "exists_ttl":     "Already exists",
        "created_ttl":    "Done",
        "created_msg":    "File created:\n{path}\n\nOpen folder?",
        "create_err":     "Failed to create file.",
        "create_err_ttl": "Error",
        "about_ttl":      "About",
        "about_ver":      "Version {ver}  |  maxsteff",
        "about_desc": (
            "Converts Jira task export files (CSV) into\n"
            "Excel spreadsheets (.xlsx).\n\n"
            "Features:\n"
            "  • Drag & drop CSV files into the window\n"
            "  • Customizable column name cleaning\n"
            "    (removes 'Custom field (…)')\n"
            "  • Column filtering via columns.txt\n"
            "  • Custom strip prefixes via [strip]\n"
            "  • Converts Jira dates → DD.MM.YYYY\n"
            "  • Excel table style selector\n"
            "  • Load multiple .csv files to merge into .xlsx\n\n"
            "columns.txt format:\n"
            "  [columns] — columns to keep\n"
            "  [strip]   — prefixes to strip\n"
            "  # lines starting with # are comments"
        ),
        "about_tg":       "Contact the author: t.me/maxsteff",
        "close":          "Close",
        "lang_ttl":       "Select language",
        "lang_auto":      "(detected automatically)",
        "apply":          "Apply",

        "s_none": "No style",
        "s_l_black": "Light — Black",
        "s_l_blue": "Light — Blue",
        "s_l_red": "Light — Red",
        "s_l_green": "Light — Green",
        "s_l_purple": "Light — Purple",
        "s_l_aqua": "Light — Aqua",
        "s_l_orange": "Light — Orange",
        "s_m_black": "Medium — Black",
        "s_m_blue": "Medium — Blue",
        "s_m_red": "Medium — Red",
        "s_m_green": "Medium — Green",
        "s_m_purple": "Medium — Purple",
        "s_m_aqua": "Medium — Aqua",
        "s_m_orange": "Medium — Orange",
        "s_d_black": "Dark — Black",
        "s_d_blue": "Dark — Blue",
        "s_d_red": "Dark — Red",
        "s_d_green": "Dark — Green",
        "s_d_purple": "Dark — Purple",
        "s_d_aqua": "Dark — Aqua",
        "s_d_orange": "Dark — Orange",
        "pick_csv_ttl":   "Select CSV file",
        "pick_xlsx_ttl":  "Save XLSX as",
        "pick_col_ttl":   "Select columns.txt",
        "csv_ft":         "CSV files",
        "xlsx_ft":        "Excel files",
        "txt_ft":         "Text files",
        "all_ft":         "All files",
        "xlsx_in_lbl":    "XLSX file (edited):",
        "csv_out_lbl":    "Save CSV as (for Jira):",
        "rh_only_lbl":    "Row height:",
        "dlm_only_lbl":   "CSV Delimiter:",
        "dlm_ttl":        "CSV Delimiter",
        "dlm_msg":        "Select a delimiter for the output file:",
        "dlm_ok":         "Continue",
        "dlm_semi":       "Semicolon ( ; )",
        "dlm_comma":      "Comma ( , )",
        "dlm_pipe":       "Pipe ( | )",
        "dlm_tab":        "Tab",
    },
    "de": {
        "title":          "Jira CSV ↔ XLSX Konverter",
        "subtitle":       "Jira tasks to excel to Jira tasks ",
        "about_btn":      "ℹ Über",
        "lang_btn":       "🌐 Sprache",
        "csv_lbl":        "CSV-Datei (Jira-Export):",
        "xlsx_lbl":       "XLSX speichern als:",
        "col_lbl":        "Spaltenfilter (columns.txt):",
        "style_lbl":      "Excel-Tabellenstil:",
        "browse":         "Durchsuchen…",
        "create_ex":      "Beispiel erstellen",
        "convert_btn":    "  ▶  Konvertieren",
        "status_idle":    "Dateien auswählen und Konvertieren klicken",
        "status_working": "Konvertierung…",
        "hint_none":      "Keine Datei — alle Spalten werden gespeichert",
        "hint_found":     "✓ Gefunden: ",
        "hint_cols":      " Spalten",
        "hint_strips":    " Strip-Präfixe",
        "hint_all":       "alle Spalten",
        "dz_idle":        "CSV-Datei hierher ziehen  (oder klicken)",
        "dz_drop":        "Datei loslassen",
        "no_file":        "Bitte CSV-Datei auswählen.",
        "no_path":        "Bitte Speicherort angeben.",
        "no_csv_ex":      "Bitte zuerst eine CSV-Datei auswählen.",
        "file_not_found": "Datei nicht gefunden:",
        "wrong_fmt":      "Bitte eine .csv-Datei auswählen",
        "wrong_fmt_ttl":  "Falsches Format",
        "done_ttl":       "Fertig!",
        "done_msg":       "Datei gespeichert:\n{path}\n\nOrdner öffnen?",
        "done_status":    "✓ Fertig: {rows} Zeilen, {cols} Spalten → {name}",
        "err_ttl":        "Konvertierungsfehler",
        "err_status":     "Fehler: {error}",
        "already_exists": "Datei existiert bereits:\n{path}",
        "exists_ttl":     "Bereits vorhanden",
        "created_ttl":    "Fertig",
        "created_msg":    "Datei erstellt:\n{path}\n\nOrdner öffnen?",
        "create_err":     "Datei konnte nicht erstellt werden.",
        "create_err_ttl": "Fehler",
        "about_ttl":      "Über",
        "about_ver":      "Version {ver}  |  maxsteff",
        "about_desc": (
            "Konvertiert Jira-Exportdateien (CSV) in\n"
            "Excel-Tabellen (.xlsx).\n\n"
            "Funktionen:\n"
            "  • Drag & Drop von CSV-Dateien in das Fenster\n"
            "  • Anpassbare Bereinigung von Spaltennamen\n"
            "    (entfernt 'Custom field (…)')\n"
            "  • Spaltenfilterung über columns.txt\n"
            "  • Benutzerdefinierte Strip-Präfixe über [strip]\n"
            "  • Konvertiert Jira-Daten → TT.MM.JJJJ\n"
            "  • Auswahl des Excel-Tabellenstils\n"
            "  • Laden mehrerer .csv-Dateien zum Zusammenführen\n\n"
            "columns.txt Format:\n"
            "  [columns] — zu behaltende Spalten\n"
            "  [strip]   — zu entfernende Präfixe\n"
            "  # Zeilen mit # sind Kommentare"
        ),
        "about_tg":       "Kontakt: t.me/maxsteff",
        "close":          "Schließen",
        "lang_ttl":       "Sprache auswählen",
        "lang_auto":      "(automatisch erkannt)",
        "apply":          "Anwenden",

        "s_none": "No style",
        "s_l_black": "Light — Black",
        "s_l_blue": "Light — Blue",
        "s_l_red": "Light — Red",
        "s_l_green": "Light — Green",
        "s_l_purple": "Light — Purple",
        "s_l_aqua": "Light — Aqua",
        "s_l_orange": "Light — Orange",
        "s_m_black": "Medium — Black",
        "s_m_blue": "Medium — Blue",
        "s_m_red": "Medium — Red",
        "s_m_green": "Medium — Green",
        "s_m_purple": "Medium — Purple",
        "s_m_aqua": "Medium — Aqua",
        "s_m_orange": "Medium — Orange",
        "s_d_black": "Dark — Black",
        "s_d_blue": "Dark — Blue",
        "s_d_red": "Dark — Red",
        "s_d_green": "Dark — Green",
        "s_d_purple": "Dark — Purple",
        "s_d_aqua": "Dark — Aqua",
        "s_d_orange": "Dark — Orange",
        "pick_csv_ttl":   "CSV-Datei auswählen",
        "pick_xlsx_ttl":  "XLSX speichern als",
        "pick_col_ttl":   "columns.txt auswählen",
        "csv_ft":         "CSV-Dateien",
        "xlsx_ft":        "Excel-Dateien",
        "txt_ft":         "Textdateien",
        "all_ft":         "Alle Dateien",
        "xlsx_in_lbl":    "XLSX-Datei (bearbeitet):",
        "csv_out_lbl":    "CSV speichern als (für Jira):",
        "rh_only_lbl":    "Zeilenhöhe:",
        "dlm_only_lbl":   "CSV-Trennzeichen:",
        "dlm_ttl":        "CSV-Trennzeichen",
        "dlm_msg":        "Wählen Sie ein Trennzeichen für die Ausgabedatei:",
        "dlm_ok":         "Weiter",
        "dlm_semi":       "Semikolon ( ; )",
        "dlm_comma":      "Komma ( , )",
        "dlm_pipe":       "Senkrechter Strich ( | )",
        "dlm_tab":        "Tabulator ( Tab )",
    },
    "fr": {
        "title":          "Convertisseur Jira CSV ↔ XLSX",
        "subtitle":       "Jira tasks to excel to Jira tasks ",
        "about_btn":      "ℹ À propos",
        "lang_btn":       "🌐 Langue",
        "csv_lbl":        "Fichier CSV (export Jira) :",
        "xlsx_lbl":       "Enregistrer XLSX sous :",
        "col_lbl":        "Filtre colonnes (columns.txt) :",
        "style_lbl":      "Style de tableau Excel :",
        "browse":         "Parcourir…",
        "create_ex":      "Créer exemple",
        "convert_btn":    "  ▶  Convertir",
        "status_idle":    "Sélectionnez les fichiers et cliquez Convertir",
        "status_working": "Conversion…",
        "hint_none":      "Aucun fichier — toutes les colonnes seront sauvées",
        "hint_found":     "✓ Trouvé : ",
        "hint_cols":      " colonnes",
        "hint_strips":    " préfixes",
        "hint_all":       "toutes colonnes",
        "dz_idle":        "Glissez un fichier .csv ici  (ou cliquez)",
        "dz_drop":        "Relâchez le fichier ici",
        "no_file":        "Veuillez sélectionner un fichier CSV.",
        "no_path":        "Veuillez indiquer où enregistrer le XLSX.",
        "no_csv_ex":      "Sélectionnez d'abord un fichier CSV.",
        "file_not_found": "Fichier introuvable :",
        "wrong_fmt":      "Veuillez sélectionner un fichier .csv",
        "wrong_fmt_ttl":  "Format incorrect",
        "done_ttl":       "Terminé !",
        "done_msg":       "Fichier enregistré :\n{path}\n\nOuvrir le dossier ?",
        "done_status":    "✓ Terminé : {rows} lignes, {cols} colonnes → {name}",
        "err_ttl":        "Erreur de conversion",
        "err_status":     "Erreur : {error}",
        "already_exists": "Le fichier existe déjà :\n{path}",
        "exists_ttl":     "Déjà existant",
        "created_ttl":    "Terminé",
        "created_msg":    "Fichier créé :\n{path}\n\nOuvrir le dossier ?",
        "create_err":     "Impossible de créer le fichier.",
        "create_err_ttl": "Erreur",
        "about_ttl":      "À propos",
        "about_ver":      "Version {ver}  |  maxsteff",
        "about_desc": (
            "Convertit les exports Jira (CSV) en\n"
            "tableaux Excel (.xlsx).\n\n"
            "Fonctionnalités :\n"
            "  • Glisser-déposer de fichiers CSV dans la fenêtre\n"
            "  • Nettoyage personnalisable des noms de colonnes\n"
            "    (supprime 'Custom field (…)')\n"
            "  • Filtrage des colonnes via columns.txt\n"
            "  • Préfixes de suppression personnalisés via [strip]\n"
            "  • Conversion des dates Jira → JJ.MM.AAAA\n"
            "  • Sélecteur de style de tableau Excel\n"
            "  • Charger plusieurs .csv pour fusionner en .xlsx\n\n"
            "Format columns.txt :\n"
            "  [columns] — colonnes à conserver\n"
            "  [strip]   — préfixes à supprimer\n"
            "  # les lignes avec # sont des commentaires"
        ),
        "about_tg":       "Contact : t.me/maxsteff",
        "close":          "Fermer",
        "lang_ttl":       "Sélectionner la langue",
        "lang_auto":      "(détecté automatiquement)",
        "apply":          "Appliquer",

        "s_none": "No style",
        "s_l_black": "Light — Black",
        "s_l_blue": "Light — Blue",
        "s_l_red": "Light — Red",
        "s_l_green": "Light — Green",
        "s_l_purple": "Light — Purple",
        "s_l_aqua": "Light — Aqua",
        "s_l_orange": "Light — Orange",
        "s_m_black": "Medium — Black",
        "s_m_blue": "Medium — Blue",
        "s_m_red": "Medium — Red",
        "s_m_green": "Medium — Green",
        "s_m_purple": "Medium — Purple",
        "s_m_aqua": "Medium — Aqua",
        "s_m_orange": "Medium — Orange",
        "s_d_black": "Dark — Black",
        "s_d_blue": "Dark — Blue",
        "s_d_red": "Dark — Red",
        "s_d_green": "Dark — Green",
        "s_d_purple": "Dark — Purple",
        "s_d_aqua": "Dark — Aqua",
        "s_d_orange": "Dark — Orange",
        "pick_csv_ttl":   "Sélectionner un fichier CSV",
        "pick_xlsx_ttl":  "Enregistrer XLSX sous",
        "pick_col_ttl":   "Sélectionner columns.txt",
        "csv_ft":         "Fichiers CSV",
        "xlsx_ft":        "Fichiers Excel",
        "txt_ft":         "Fichiers texte",
        "all_ft":         "Tous les fichiers",
        "xlsx_in_lbl":    "Fichier XLSX (édité) :",
        "csv_out_lbl":    "Enregistrer CSV sous (pour Jira) :",
        "rh_only_lbl":    "Hauteur de ligne :",
        "dlm_only_lbl":   "Séparateur CSV :",
        "dlm_ttl":        "Séparateur CSV",
        "dlm_msg":        "Sélectionnez un séparateur pour le fichier de sortie :",
        "dlm_ok":         "Continuer",
        "dlm_semi":       "Point-virgule ( ; )",
        "dlm_comma":      "Virgule ( , )",
        "dlm_pipe":       "Barre verticale ( | )",
        "dlm_tab":        "Tabulation ( Tab )",
    },
    "es": {
        "title":          "Convertidor Jira CSV ↔ XLSX",
        "subtitle":       "Jira tasks to excel to Jira tasks ",
        "about_btn":      "ℹ Acerca de",
        "lang_btn":       "🌐 Idioma",
        "csv_lbl":        "Archivo CSV (exportación Jira):",
        "xlsx_lbl":       "Guardar XLSX como:",
        "col_lbl":        "Filtro columnas (columns.txt):",
        "style_lbl":      "Estilo de tabla Excel:",
        "browse":         "Explorar…",
        "create_ex":      "Crear ejemplo",
        "convert_btn":    "  ▶  Convertir",
        "status_idle":    "Seleccione archivos y haga clic en Convertir",
        "status_working": "Convirtiendo…",
        "hint_none":      "Sin archivo — se guardarán todas las columnas",
        "hint_found":     "✓ Encontrado: ",
        "hint_cols":      " columnas",
        "hint_strips":    " prefijos",
        "hint_all":       "todas las columnas",
        "dz_idle":        "Arrastre un archivo .csv aquí  (o haga clic)",
        "dz_drop":        "Suelte el archivo aquí",
        "no_file":        "Seleccione un archivo CSV.",
        "no_path":        "Indique dónde guardar el XLSX.",
        "no_csv_ex":      "Primero seleccione un archivo CSV.",
        "file_not_found": "Archivo no encontrado:",
        "wrong_fmt":      "Seleccione un archivo .csv",
        "wrong_fmt_ttl":  "Formato incorrecto",
        "done_ttl":       "¡Listo!",
        "done_msg":       "Archivo guardado:\n{path}\n\n¿Abrir carpeta?",
        "done_status":    "✓ Listo: {rows} filas, {cols} columnas → {name}",
        "err_ttl":        "Error de conversión",
        "err_status":     "Error: {error}",
        "already_exists": "El archivo ya existe:\n{path}",
        "exists_ttl":     "Ya existe",
        "created_ttl":    "Listo",
        "created_msg":    "Archivo creado:\n{path}\n\n¿Abrir carpeta?",
        "create_err":     "No se pudo crear el archivo.",
        "create_err_ttl": "Error",
        "about_ttl":      "Acerca de",
        "about_ver":      "Versión {ver}  |  maxsteff",
        "about_desc": (
            "Convierte exportaciones de Jira (CSV) a\n"
            "hojas de cálculo Excel (.xlsx).\n\n"
            "Funciones:\n"
            "  • Arrastrar y soltar archivos CSV en la ventana\n"
            "  • Limpieza personalizable de nombres de columnas\n"
            "    (elimina 'Custom field (…)')\n"
            "  • Filtrado de columnas mediante columns.txt\n"
            "  • Prefijos de recorte personalizados en [strip]\n"
            "  • Convierte fechas de Jira → DD.MM.AAAA\n"
            "  • Selector de estilo de tabla de Excel\n"
            "  • Cargar varios .csv para combinarlos en .xlsx\n\n"
            "Formato columns.txt:\n"
            "  [columns] — columnas a mantener\n"
            "  [strip]   — prefijos a recortar\n"
            "  # las líneas con # son comentarios"
        ),
        "about_tg":       "Contacto: t.me/maxsteff",
        "close":          "Cerrar",
        "lang_ttl":       "Seleccionar idioma",
        "lang_auto":      "(detectado automáticamente)",
        "apply":          "Aplicar",

        "s_none": "No style",
        "s_l_black": "Light — Black",
        "s_l_blue": "Light — Blue",
        "s_l_red": "Light — Red",
        "s_l_green": "Light — Green",
        "s_l_purple": "Light — Purple",
        "s_l_aqua": "Light — Aqua",
        "s_l_orange": "Light — Orange",
        "s_m_black": "Medium — Black",
        "s_m_blue": "Medium — Blue",
        "s_m_red": "Medium — Red",
        "s_m_green": "Medium — Green",
        "s_m_purple": "Medium — Purple",
        "s_m_aqua": "Medium — Aqua",
        "s_m_orange": "Medium — Orange",
        "s_d_black": "Dark — Black",
        "s_d_blue": "Dark — Blue",
        "s_d_red": "Dark — Red",
        "s_d_green": "Dark — Green",
        "s_d_purple": "Dark — Purple",
        "s_d_aqua": "Dark — Aqua",
        "s_d_orange": "Dark — Orange",
        "pick_csv_ttl":   "Seleccionar archivo CSV",
        "pick_xlsx_ttl":  "Guardar XLSX como",
        "pick_col_ttl":   "Seleccionar columns.txt",
        "csv_ft":         "Archivos CSV",
        "xlsx_ft":        "Archivos Excel",
        "txt_ft":         "Archivos de texto",
        "all_ft":         "Todos los archivos",
        "xlsx_in_lbl":    "Archivo XLSX (editado):",
        "csv_out_lbl":    "Guardar CSV como (para Jira):",
        "rh_only_lbl":    "Altura de fila:",
        "dlm_only_lbl":   "Separador CSV:",
        "dlm_ttl":        "Separador CSV",
        "dlm_msg":        "Seleccione un separador para el archivo de salida:",
        "dlm_ok":         "Continuar",
        "dlm_semi":       "Punto y coma ( ; )",
        "dlm_comma":      "Coma ( , )",
        "dlm_pipe":       "Barra vertical ( | )",
        "dlm_tab":        "Tabulador ( Tab )",
    },
    "zh": {
        "title":          "Jira CSV ↔ XLSX 转换器",
        "subtitle":       "Jira tasks to excel to Jira tasks ",
        "about_btn":      "ℹ 关于",
        "lang_btn":       "🌐 语言",
        "csv_lbl":        "CSV 文件（Jira 导出）：",
        "xlsx_lbl":       "保存 XLSX 为：",
        "col_lbl":        "列过滤器 (columns.txt)：",
        "style_lbl":      "Excel 表格样式：",
        "browse":         "浏览…",
        "create_ex":      "创建示例",
        "convert_btn":    "  ▶  转换",
        "status_idle":    "选择文件后点击转换",
        "status_working": "转换中…",
        "hint_none":      "未选择文件 — 将保存所有列",
        "hint_found":     "✓ 已找到：",
        "hint_cols":      " 列",
        "hint_strips":    " 前缀",
        "hint_all":       "所有列",
        "dz_idle":        "将 .csv 文件拖到此处（或点击）",
        "dz_drop":        "在此释放文件",
        "no_file":        "请选择 CSV 文件。",
        "no_path":        "请指定 XLSX 保存位置。",
        "no_csv_ex":      "请先选择 CSV 文件。",
        "file_not_found": "文件未找到：",
        "wrong_fmt":      "请选择 .csv 文件",
        "wrong_fmt_ttl":  "格式错误",
        "done_ttl":       "完成！",
        "done_msg":       "文件已保存：\n{path}\n\n打开文件夹？",
        "done_status":    "✓ 完成：{rows} 行，{cols} 列 → {name}",
        "err_ttl":        "转换错误",
        "err_status":     "错误：{error}",
        "already_exists": "文件已存在：\n{path}",
        "exists_ttl":     "已存在",
        "created_ttl":    "完成",
        "created_msg":    "文件已创建：\n{path}\n\n打开文件夹？",
        "create_err":     "无法创建文件。",
        "create_err_ttl": "错误",
        "about_ttl":      "关于",
        "about_ver":      "版本 {ver}  |  maxsteff",
        "about_desc": (
            "将 Jira 导出的 CSV 文件转换为\n"
            "Excel 表格 (.xlsx)。\n\n"
            "功能：\n"
            "  • 将 CSV 文件拖放到窗口中\n"
            "  • 可自定义列名清理\n"
            "    (删除 'Custom field (…)')\n"
            "  • 通过 columns.txt 过滤列\n"
            "  • 通过 [strip] 自定义去除前缀\n"
            "  • 转换 Jira 日期 → DD.MM.YYYY\n"
            "  • Excel 表格样式选择器\n"
            "  • 加载多个 .csv 文件以合并为 .xlsx\n\n"
            "columns.txt 格式：\n"
            "  [columns] — 要保留的列\n"
            "  [strip]   — 要去除的前缀\n"
            "  # 开头带有 # 的行是注释"
        ),
        "about_tg":       "联系作者：t.me/maxsteff",
        "close":          "关闭",
        "lang_ttl":       "选择语言",
        "lang_auto":      "（自动检测）",
        "apply":          "应用",

        "s_none": "No style",
        "s_l_black": "Light — Black",
        "s_l_blue": "Light — Blue",
        "s_l_red": "Light — Red",
        "s_l_green": "Light — Green",
        "s_l_purple": "Light — Purple",
        "s_l_aqua": "Light — Aqua",
        "s_l_orange": "Light — Orange",
        "s_m_black": "Medium — Black",
        "s_m_blue": "Medium — Blue",
        "s_m_red": "Medium — Red",
        "s_m_green": "Medium — Green",
        "s_m_purple": "Medium — Purple",
        "s_m_aqua": "Medium — Aqua",
        "s_m_orange": "Medium — Orange",
        "s_d_black": "Dark — Black",
        "s_d_blue": "Dark — Blue",
        "s_d_red": "Dark — Red",
        "s_d_green": "Dark — Green",
        "s_d_purple": "Dark — Purple",
        "s_d_aqua": "Dark — Aqua",
        "s_d_orange": "Dark — Orange",
        "pick_csv_ttl":   "选择 CSV 文件",
        "pick_xlsx_ttl":  "保存 XLSX 为",
        "pick_col_ttl":   "选择 columns.txt",
        "csv_ft":         "CSV 文件",
        "xlsx_ft":        "Excel 文件",
        "txt_ft":         "文本文件",
        "all_ft":         "所有文件",
        "xlsx_in_lbl":    "XLSX 文件（编辑后）：",
        "csv_out_lbl":    "保存 CSV 为（用于 Jira）：",
        "rh_only_lbl":    "行高：",
        "dlm_only_lbl":   "CSV 分隔符：",
        "dlm_ttl":        "CSV 分隔符",
        "dlm_msg":        "为输出文件选择一个分隔符：",
        "dlm_ok":         "继续",
        "dlm_semi":       "分号 ( ; )",
        "dlm_comma":      "逗号 ( , )",
        "dlm_pipe":       "竖线 ( | )",
        "dlm_tab":        "制表符 ( Tab )",
    },
    "ar": {
        "title":          "محوّل Jira CSV ↔ XLSX",
        "subtitle":       "Jira tasks to excel to Jira tasks ",
        "about_btn":      "ℹ حول البرنامج",
        "lang_btn":       "🌐 اللغة",
        "csv_lbl":        "ملف CSV (تصدير Jira):",
        "xlsx_lbl":       "حفظ XLSX كـ:",
        "col_lbl":        "فلتر الأعمدة (columns.txt):",
        "style_lbl":      "نمط جدول Excel:",
        "browse":         "تصفح…",
        "create_ex":      "إنشاء مثال",
        "convert_btn":    "  ▶  تحويل",
        "status_idle":    "اختر الملفات ثم اضغط تحويل",
        "status_working": "جارٍ التحويل…",
        "hint_none":      "لم يُحدَّد ملف — ستُحفظ جميع الأعمدة",
        "hint_found":     "✓ تم العثور: ",
        "hint_cols":      " أعمدة",
        "hint_strips":    " بادئات",
        "hint_all":       "جميع الأعمدة",
        "dz_idle":        "اسحب ملف .csv هنا  (أو انقر)",
        "dz_drop":        "أفلت الملف هنا",
        "no_file":        "الرجاء تحديد ملف CSV.",
        "no_path":        "الرجاء تحديد مكان الحفظ.",
        "no_csv_ex":      "الرجاء تحديد ملف CSV أولاً.",
        "file_not_found": "الملف غير موجود:",
        "wrong_fmt":      "الرجاء تحديد ملف .csv",
        "wrong_fmt_ttl":  "تنسيق خاطئ",
        "done_ttl":       "تم!",
        "done_msg":       "تم حفظ الملف:\n{path}\n\nفتح المجلد؟",
        "done_status":    "✓ تم: {rows} صفوف، {cols} أعمدة → {name}",
        "err_ttl":        "خطأ في التحويل",
        "err_status":     "خطأ: {error}",
        "already_exists": "الملف موجود بالفعل:\n{path}",
        "exists_ttl":     "موجود بالفعل",
        "created_ttl":    "تم",
        "created_msg":    "تم إنشاء الملف:\n{path}\n\nفتح المجلد؟",
        "create_err":     "فشل إنشاء الملف.",
        "create_err_ttl": "خطأ",
        "about_ttl":      "حول البرنامج",
        "about_ver":      "الإصدار {ver}  |  maxsteff",
        "about_desc": (
            "يحوّل ملفات تصدير Jira (CSV) إلى\n"
            "جداول Excel (.xlsx).\n\n"
            "الميزات:\n"
            "  • سحب وإفلات ملفات CSV في النافذة\n"
            "  • تنظيف أسماء الأعمدة القابل للتخصيص\n"
            "    (يزيل 'Custom field (…)')\n"
            "  • تصفية الأعمدة عبر ملف columns.txt\n"
            "  • بادئات إزالة مخصصة عبر [strip]\n"
            "  • تحويل تواريخ Jira → يوم.شهر.سنة\n"
            "  • اختيار نمط جدول Excel\n"
            "  • تحميل عدة ملفات .csv لدمجها في .xlsx\n\n"
            "تنسيق columns.txt:\n"
            "  [columns] — الأعمدة المراد الاحتفاظ بها\n"
            "  [strip]   — البادئات المراد إزالتها\n"
            "  # الأسطر التي تبدأ بـ # هي تعليقات"
        ),
        "about_tg":       "تواصل مع المؤلف: t.me/maxsteff",
        "close":          "إغلاق",
        "lang_ttl":       "اختر اللغة",
        "lang_auto":      "(تم الكشف تلقائياً)",
        "apply":          "تطبيق",

        "s_none": "No style",
        "s_l_black": "Light — Black",
        "s_l_blue": "Light — Blue",
        "s_l_red": "Light — Red",
        "s_l_green": "Light — Green",
        "s_l_purple": "Light — Purple",
        "s_l_aqua": "Light — Aqua",
        "s_l_orange": "Light — Orange",
        "s_m_black": "Medium — Black",
        "s_m_blue": "Medium — Blue",
        "s_m_red": "Medium — Red",
        "s_m_green": "Medium — Green",
        "s_m_purple": "Medium — Purple",
        "s_m_aqua": "Medium — Aqua",
        "s_m_orange": "Medium — Orange",
        "s_d_black": "Dark — Black",
        "s_d_blue": "Dark — Blue",
        "s_d_red": "Dark — Red",
        "s_d_green": "Dark — Green",
        "s_d_purple": "Dark — Purple",
        "s_d_aqua": "Dark — Aqua",
        "s_d_orange": "Dark — Orange",
        "pick_csv_ttl":   "اختر ملف CSV",
        "pick_xlsx_ttl":  "حفظ XLSX كـ",
        "pick_col_ttl":   "اختر columns.txt",
        "csv_ft":         "ملفات CSV",
        "xlsx_ft":        "ملفات Excel",
        "txt_ft":         "ملفات نصية",
        "all_ft":         "جميع الملفات",
        "xlsx_in_lbl":    "ملف XLSX (مُعدّل):",
        "csv_out_lbl":    "حفظ CSV كـ (لـ Jira):",
        "rh_only_lbl":    "ارتفاع الصف:",
        "dlm_only_lbl":   "فاصل CSV:",
        "dlm_ttl":        "فاصل CSV",
        "dlm_msg":        "حدد فاصلاً لملف الإخراج:",
        "dlm_ok":         "متابعة",
        "dlm_semi":       "فاصلة منقوطة ( ; )",
        "dlm_comma":      "فاصلة ( , )",
        "dlm_pipe":       "خط عمودي ( | )",
        "dlm_tab":        "علامة جدولة ( Tab )",
    },
    "pt": {
        "title":          "Conversor Jira CSV ↔ XLSX",
        "subtitle":       "Jira tasks to excel to Jira tasks ",
        "about_btn":      "ℹ Sobre",
        "lang_btn":       "🌐 Idioma",
        "csv_lbl":        "Ficheiro CSV (exportação Jira):",
        "xlsx_lbl":       "Guardar XLSX como:",
        "col_lbl":        "Filtro colunas (columns.txt):",
        "style_lbl":      "Estilo de tabela Excel:",
        "browse":         "Procurar…",
        "create_ex":      "Criar exemplo",
        "convert_btn":    "  ▶  Converter",
        "status_idle":    "Selecione os ficheiros e clique em Converter",
        "status_working": "A converter…",
        "hint_none":      "Nenhum ficheiro — todas as colunas serão guardadas",
        "hint_found":     "✓ Encontrado: ",
        "hint_cols":      " colunas",
        "hint_strips":    " prefixos",
        "hint_all":       "todas as colunas",
        "dz_idle":        "Arraste um ficheiro .csv aqui  (ou clique)",
        "dz_drop":        "Largue o ficheiro aqui",
        "no_file":        "Selecione um ficheiro CSV.",
        "no_path":        "Indique onde guardar o XLSX.",
        "no_csv_ex":      "Selecione primeiro um ficheiro CSV.",
        "file_not_found": "Ficheiro não encontrado:",
        "wrong_fmt":      "Selecione um ficheiro .csv",
        "wrong_fmt_ttl":  "Formato incorreto",
        "done_ttl":       "Concluído!",
        "done_msg":       "Ficheiro guardado:\n{path}\n\nAbrir pasta?",
        "done_status":    "✓ Concluído: {rows} linhas, {cols} colunas → {name}",
        "err_ttl":        "Erro de conversão",
        "err_status":     "Erro: {error}",
        "already_exists": "O ficheiro já existe:\n{path}",
        "exists_ttl":     "Já existe",
        "created_ttl":    "Concluído",
        "created_msg":    "Ficheiro criado:\n{path}\n\nAbrir pasta?",
        "create_err":     "Não foi possível criar o ficheiro.",
        "create_err_ttl": "Erro",
        "about_ttl":      "Sobre",
        "about_ver":      "Versão {ver}  |  maxsteff",
        "about_desc": (
            "Converte exportações do Jira (CSV) em\n"
            "folhas de cálculo Excel (.xlsx).\n\n"
            "Funcionalidades:\n"
            "  • Arrastar e soltar ficheiros CSV na janela\n"
            "  • Limpeza personalizável de nomes de colunas\n"
            "    (remove 'Custom field (…)')\n"
            "  • Filtragem de colunas via columns.txt\n"
            "  • Prefixos de corte personalizados em [strip]\n"
            "  • Converte datas do Jira → DD.MM.AAAA\n"
            "  • Seletor de estilo de tabela Excel\n"
            "  • Carregar múltiplos .csv para unir num .xlsx\n\n"
            "Formato columns.txt:\n"
            "  [columns] — colunas a manter\n"
            "  [strip]   — prefixos a cortar\n"
            "  # linhas com # são comentários"
        ),
        "about_tg":       "Contacto: t.me/maxsteff",
        "close":          "Fechar",
        "lang_ttl":       "Selecionar idioma",
        "lang_auto":      "(detetado automaticamente)",
        "apply":          "Aplicar",

        "s_none": "No style",
        "s_l_black": "Light — Black",
        "s_l_blue": "Light — Blue",
        "s_l_red": "Light — Red",
        "s_l_green": "Light — Green",
        "s_l_purple": "Light — Purple",
        "s_l_aqua": "Light — Aqua",
        "s_l_orange": "Light — Orange",
        "s_m_black": "Medium — Black",
        "s_m_blue": "Medium — Blue",
        "s_m_red": "Medium — Red",
        "s_m_green": "Medium — Green",
        "s_m_purple": "Medium — Purple",
        "s_m_aqua": "Medium — Aqua",
        "s_m_orange": "Medium — Orange",
        "s_d_black": "Dark — Black",
        "s_d_blue": "Dark — Blue",
        "s_d_red": "Dark — Red",
        "s_d_green": "Dark — Green",
        "s_d_purple": "Dark — Purple",
        "s_d_aqua": "Dark — Aqua",
        "s_d_orange": "Dark — Orange",
        "pick_csv_ttl":   "Selecionar ficheiro CSV",
        "pick_xlsx_ttl":  "Guardar XLSX como",
        "pick_col_ttl":   "Selecionar columns.txt",
        "csv_ft":         "Ficheiros CSV",
        "xlsx_ft":        "Ficheiros Excel",
        "txt_ft":         "Ficheiros de texto",
        "all_ft":         "Todos os ficheiros",
        "xlsx_in_lbl":    "Ficheiro XLSX (editado):",
        "csv_out_lbl":    "Guardar CSV como (para Jira):",
        "rh_only_lbl":    "Altura da linha:",
        "dlm_only_lbl":   "Separador CSV:",
        "dlm_ttl":        "Separador CSV",
        "dlm_msg":        "Selecione um separador para o ficheiro de saída:",
        "dlm_ok":         "Continuar",
        "dlm_semi":       "Ponto e vírgula ( ; )",
        "dlm_comma":      "Vírgula ( , )",
        "dlm_pipe":       "Barra vertical ( | )",
        "dlm_tab":        "Tabulação ( Tab )",
    },
}


def detect_system_language() -> str:
    """
    Определяет код языка системы (например 'ru', 'en', 'de').
    Возвращает 'en' если не получилось определить или язык не поддерживается.

    Источники проверяются по очереди:
      1. locale.getdefaultlocale() — обычно срабатывает на всех ОС
      2. реестр Windows (HKCU\\Control Panel\\International\\LocaleName)
    """
    try:
        import locale
        lang = locale.getdefaultlocale()[0] or ""
        code = lang.split("_")[0].lower()
        if code in STRINGS:
            return code
    except Exception:
        pass
    try:
        key = winreg.OpenKey(winreg.HKEY_CURRENT_USER,
            r"Control Panel\International")
        val, _ = winreg.QueryValueEx(key, "LocaleName")
        winreg.CloseKey(key)
        code = val.split("-")[0].lower()
        if code in STRINGS:
            return code
    except Exception:
        pass
    return "en"


def detect_system_theme() -> str:
    """
    Определяет режим Windows (светлый/тёмный) через реестр.
    AppsUseLightTheme = 1 → светлая, 0 → тёмная.
    Если ключ недоступен (старая Windows, ошибка), возвращает 'light'.
    """
    try:
        key = winreg.OpenKey(winreg.HKEY_CURRENT_USER,
            r"Software\Microsoft\Windows\CurrentVersion\Themes\Personalize")
        val, _ = winreg.QueryValueEx(key, "AppsUseLightTheme")
        winreg.CloseKey(key)
        return "light" if val == 1 else "dark"
    except Exception:
        return "light"


# ─── Файловая система ────────────────────────────────────────────────────────

def get_exe_dir() -> Path:
    """
    Папка с исполняемым файлом — место для columns.txt и settings.json.
    В режиме PyInstaller (sys.frozen=True) это папка с .exe, иначе — папка
    со скриптом. ВАЖНО: sys._MEIPASS использовать НЕЛЬЗЯ — это временная
    папка распаковки onefile-сборки, которая создаётся при каждом запуске.
    """
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent

def resource_path(relative_path):
    """Получает абсолютный путь к ресурсу (работает и в dev, и в собранном .exe)"""
    try:
        # PyInstaller создает временную папку и хранит путь в _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

def settings_path() -> Path:
    return get_exe_dir() / SETTINGS_FILE


def load_settings() -> dict:
    """Читает settings.json. Молча возвращает {} при любых проблемах."""
    p = settings_path()
    if p.exists():
        try:
            return json.loads(p.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {}


def save_settings(data: dict):
    """Пишет settings.json. Ошибки молча проглатываются — не критично."""
    try:
        settings_path().write_text(json.dumps(data, ensure_ascii=False, indent=2),
                                   encoding="utf-8")
    except Exception:
        pass


def find_columns_txt() -> str:
    """
    Ищет columns.txt по типичным путям (рядом с .exe, в рабочей директории,
    рядом со скриптом). Возвращает путь или пустую строку.
    """
    candidates = [Path(sys.executable).resolve().parent, Path(os.getcwd()).resolve()]
    if not getattr(sys, "frozen", False):
        candidates.append(Path(__file__).resolve().parent)
    for p in candidates:
        f = p / "columns.txt"
        if f.exists():
            return str(f)
    return ""


def find_csv_in_exe_dir() -> str:
    """Первый .csv рядом с .exe (для автоподстановки при запуске) или ''."""
    csv_files = sorted(get_exe_dir().glob("*.csv"))
    return str(csv_files[0]) if csv_files else ""


# ─── Разбор columns.txt ──────────────────────────────────────────────────────

def parse_columns_txt(path: str) -> dict:
    """
    Разбирает columns.txt с поддержкой секций.

    Формат:
        [columns]                    ← по умолчанию активна эта секция
        Ключ проблемы                ← список колонок для сохранения,
        Статус                          порядок сохраняется в xlsx
        ...
        [strip]
        Пользовательское поле        ← префиксы которые нужно срезать
        Custom field                    из имён колонок CSV

    Строки с # — комментарии. Пустые строки игнорируются.

    Тонкость: в [strip] пробелы в конце строки СОХРАНЯЮТСЯ — они часть
    префикса. "Пользовательское поле " (с пробелом) при срезке оставит
    "Тема" из "Пользовательское поле Тема", без пробела — "(Тема)".
    """
    result = {"columns": [], "strip_prefixes": []}
    if not path or not Path(path).exists():
        return result

    current = "columns"   # секция по умолчанию для совместимости со старым форматом
    for raw in Path(path).read_text(encoding="utf-8-sig").splitlines():
        stripped = raw.strip()
        if not stripped or stripped.startswith("#"):
            continue
        if stripped.lower() == "[columns]":
            current = "columns"
            continue
        if stripped.lower() == "[strip]":
            current = "strip_prefixes"
            continue
        # В [strip] нельзя trim() — пробелы значимы
        value = raw.rstrip("\n\r") if current == "strip_prefixes" else stripped
        result[current].append(value)
    return result


def auto_rename(col: str, strip_prefixes: list | None = None) -> str:
    """
    Очищает имя колонки от служебных обёрток.

    Логика:
      1. Если задан список префиксов из [strip] — пробуем их по очереди.
         Распознаются два формата:  "PREFIX(name)"  или  "PREFIX name".
      2. Fallback — встроенный шаблон "Пользовательское поле (...)".
      3. Если ничего не подошло — имя возвращается как есть.
    """
    stripped = col.strip()

    if strip_prefixes:
        for prefix in strip_prefixes:
            # Вариант 1: префикс + (имя в скобках)
            pat = re.match(re.escape(prefix) + r"\s*\((.+)\)\s*$", stripped)
            if pat:
                return pat.group(1).strip()
            # Вариант 2: префикс сразу за которым идёт имя
            if stripped.startswith(prefix):
                return stripped[len(prefix):].strip().strip("()")

    # Встроенный шаблон Jira
    m = re.match(r"Пользовательское поле \((.+)\)$", stripped)
    return m.group(1).strip() if m else stripped


def load_columns_filter(explicit_path: str = "") -> tuple:
    """
    Возвращает (keep_list, strip_prefixes) для использования в convert().
      keep_list      — список колонок для фильтрации, или None если фильтр пустой
      strip_prefixes — список префиксов для auto_rename()

    Если explicit_path задан и существует — берём его, иначе автопоиск.
    """
    if explicit_path and Path(explicit_path).exists():
        cfg_path = explicit_path
    else:
        cfg_path = find_columns_txt()

    if not cfg_path:
        return None, []

    parsed = parse_columns_txt(cfg_path)
    cols   = parsed["columns"] or None   # пустой список → None (фильтра нет)
    strips = parsed["strip_prefixes"]
    return cols, strips


# ─── Разбор дат Jira ─────────────────────────────────────────────────────────

def parse_jira_date(value):
    """
    Конвертирует "05/фев/26 12:00 AM" → "05.02.2026".
    Время отбрасывается, двузначный год дополняется до "20XX".
    Не-строки и нераспознанные значения возвращаются без изменений.
    """
    if not isinstance(value, str):
        return value
    m = re.match(r"(\d{1,2})/([а-яёА-ЯЁa-zA-Z]+)/(\d{2,4})", value.strip())
    if m:
        day, mon, year = m.group(1), m.group(2).lower(), m.group(3)
        mn = MONTH_MAP.get(mon[:3])
        if mn:
            if len(year) == 2:
                year = "20" + year
            return f"{day.zfill(2)}.{mn}.{year}"
    return value


def generate_columns_example(csv_path: str, save_dir: Path,
                              strip_prefixes: list | None = None,
                              delimiter: str = ";") -> str:
    """
    Создаёт columns.txt-шаблон на основе заголовков переданного CSV.
    Все колонки попадают в секцию [columns] с применённым auto_rename().
    Возвращает путь к созданному файлу или '' при ошибке.
    """
    out = save_dir / "columns.txt"
    if delimiter.lower() == "tab": delimiter = "\t"
    try:
        # Читаем только заголовки — данные не нужны, экономим память
        df = pd.read_csv(csv_path, sep=delimiter, encoding="utf-8-sig", dtype=str, nrows=0)
        renamed = [auto_rename(c, strip_prefixes) for c in df.columns]
        lines = [
            "# ──────────────────────────────────────────────────────────────────────",
            "# Пример фильтра колонок — сгенерирован автоматически из CSV.",
            "# Скопируйте этот файл как columns.txt и удалите ненужные строки.",
            "# Строки начинающиеся с # — комментарии, игнорируются.",
            "#",
            "# СЕКЦИИ:",
            "#   [columns]  — список колонок для сохранения (по умолчанию)",
            "#   [strip]    — префиксы которые нужно обрезать из названий",
            "#",
            "# Пример секции [strip]:",
            "#   [strip]",
            "#   Пользовательское поле ",
            "#   Custom field ",
            "# ──────────────────────────────────────────────────────────────────────",
            "",
            "[columns]",
            "",
        ] + renamed + [
            "",
            "# [strip]",
            "# Пользовательское поле ",
        ]
        out.write_text("\n".join(lines), encoding="utf-8")
        return str(out)
    except Exception:
        return ""


# ─── Главная функция конвертации ─────────────────────────────────────────────

def convert(csv_paths: list, xlsx_path: str,
            columns_path: str = "",
            table_style: str | None = None,
            row_height: str = "Авто",
            delimiter: str = ";") -> tuple:
    """
    CSV → XLSX. Возвращает (количество_строк, количество_колонок).

    Параметры:
        csv_path     — исходный CSV (разделитель ';', UTF-8 BOM)
        xlsx_path    — куда сохранить
        columns_path — путь к columns.txt (или '' для автопоиска)
        table_style  — имя встроенного стиля Excel или None

    Этапы:
      1. Читаем CSV, переименовываем колонки через auto_rename()
      2. Если в [columns] задан фильтр — оставляем только нужные
      3. Колонки с "дата" в названии прогоняем через parse_jira_date()
      4. Записываем в xlsx, применяем выравнивание/границы/ширину
      5. Если задан table_style — оборачиваем диапазон в Table
         (Excel сам отрисует заголовок и полосы по выбранному стилю)
    """    
    keep, strip_prefixes = load_columns_filter(columns_path)
    if delimiter.lower() == "tab": delimiter = "\t"
    # Читаем все переданные CSV-файлы и склеиваем в один DataFrame
    dfs = []
    for p in csv_paths:
        dfs.append(pd.read_csv(p, sep=delimiter, encoding="utf-8-sig", dtype=str))
        
    if not dfs:
        return 0, 0
        
    df = pd.concat(dfs, ignore_index=True)
    df = df.drop_duplicates(ignore_index=True)
    
    # Удаляем дубликаты задач (на случай, если пользователь загрузил одни и те же файлы)
    df = df.drop_duplicates(ignore_index=True)

    df = df.rename(columns={c: auto_rename(c, strip_prefixes) for c in df.columns})

    # Фильтр колонок: сохраняем только запрошенные, в указанном порядке
    if keep:
        df = df[[c for c in keep if c in df.columns]]

    # Конвертация дат — ищем колонки по подстроке "дата" в имени
    for col in [c for c in df.columns if "дата" in c.lower()]:
        df[col] = df[col].apply(parse_jira_date)

    wb = Workbook()
    ws = wb.active
    ws.title = "Tasks"

    # Заголовок и данные
    for ci, name in enumerate(df.columns, 1):
        ws.cell(row=1, column=ci, value=name)
    for ri, row in enumerate(df.itertuples(index=False), 2):
        for ci, val in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=("" if pd.isna(val) else val))

    # Оформление ячеек данных: перенос строк + тонкая серая граница
    ca   = Alignment(vertical="top", wrap_text=True)
    thin = Side(style="thin", color="BDC3C7")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ri in range(2, len(df) + 2):
        # --- Установка высоты строки ---
        if str(row_height).strip().lower() not in ["авто", "auto", ""]:
            try:
                ws.row_dimensions[ri].height = float(row_height)
            except ValueError:
                pass # Если ввели текст по ошибке, оставляем авто-высоту
        # -------------------------------
        for ci in range(1, len(df.columns) + 1):
            cell = ws.cell(row=ri, column=ci)
            cell.alignment = ca
            cell.border = brd

    # Авто-ширина колонок: ~85% от длины самого длинного значения,
    # но в пределах [12, 50] символов
    for ci, col in enumerate(df.columns, 1):
        col_data = df.iloc[:, ci - 1].astype(str)
        mx = max(len(str(col)), col_data.str.len().max() if len(df) else 0)
        ws.column_dimensions[get_column_letter(ci)].width = min(max(mx * 0.85, 12), 50)

    ws.freeze_panes = "A2"   # закрепляем строку заголовка

    # Стиль таблицы Excel
    if table_style:
        # Оборачиваем диапазон в Table — Excel сам отрисует заголовок и полосы,
        # автофильтр включается автоматически
        last_col = get_column_letter(len(df.columns))
        ref = f"A1:{last_col}{len(df) + 1}"
        tbl = Table(displayName="DataTable", ref=ref)
        tbl.tableStyleInfo = TableStyleInfo(
            name=table_style,
            showFirstColumn=False, showLastColumn=False,
            showRowStripes=True, showColumnStripes=False)
        ws.add_table(tbl)
    else:
        # Без стиля — рисуем простую строгую таблицу:
        # Чёрные границы для всех ячеек, жирный чёрный текст по центру для заголовка
        thin_black = Side(style="thin", color="000000")
        brd_black  = Border(left=thin_black, right=thin_black, top=thin_black, bottom=thin_black)
        
        hf = Font(name="Calibri", bold=True, color="000000", size=10)
        ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
        # Оформляем заголовок (строка 1)
        for ci in range(1, len(df.columns) + 1):
            cell = ws.cell(row=1, column=ci)
            cell.font = hf
            cell.alignment = ha
            cell.border = brd_black
        # Оформляем данные (чёрные рамки со 2 строки)
        for ri in range(2, len(df) + 2):
            # --- Установка высоты строки ---
            if str(row_height).strip().lower() not in ["авто", "auto", ""]:
                try:
                    ws.row_dimensions[ri].height = float(row_height)
                except ValueError:
                    pass # Если ввели текст по ошибке, оставляем авто-высоту
            # -------------------------------
            for ci in range(1, len(df.columns) + 1):
                ws.cell(row=ri, column=ci).border = brd_black
        ws.auto_filter.ref = ws.dimensions   # автофильтр в "ручном" режиме

    ws.row_dimensions[1].height = 36   # высота заголовка под перенос длинных имён
    wb.save(xlsx_path)
    return len(df), len(df.columns)


# ═════════════════════════════════════════════════════════════════════════════
# GUI
# ═════════════════════════════════════════════════════════════════════════════
def convert_xlsx_to_csv(xlsx_paths: list, csv_path: str, delimiter: str = ";") -> tuple:
    """XLSX → CSV для обратной загрузки в Jira."""
    if delimiter.lower() == "tab": delimiter = "\t"
    dfs = []
    for p in xlsx_paths:
        dfs.append(pd.read_excel(p, dtype=str))
        
    if not dfs:
        return 0, 0
        
    df = pd.concat(dfs, ignore_index=True)
    # ИСПРАВЛЕНО: используем чистый utf-8
    df.to_csv(csv_path, sep=delimiter, encoding="utf-8", index=False)
    return len(df), len(df.columns)

class App(TkinterDnD.Tk if _DND else tk.Tk):
    """
    Главное окно приложения. Наследуется от TkinterDnD.Tk если установлен
    tkinterdnd2 (для drag & drop файлов), иначе от обычного tk.Tk.

    Структура:
      • header     — синяя шапка с кнопками "О программе" / "Язык" / тема
      • drop_zone  — пунктирная область для перетаскивания CSV
      • card       — карточка с тремя полями (CSV, XLSX, columns.txt)
                     и селектором стиля таблицы
      • status     — строка статуса под карточкой
      • progress   — индикатор прогресса конвертации
    """

    def __init__(self):
        super().__init__()
        self.title(APP_NAME)
        self.resizable(True, False)

        # --- ДОБАВЛЕННЫЙ КОД ДЛЯ ИКОНКИ ---
        try:
            self.iconbitmap(resource_path(os.path.join("assets", "icon.ico")))
        except Exception as e:
            pass # Если иконки нет (например, при запуске скрипта без папки assets), программа не упадет
        # ----------------------------------
        self.title(APP_NAME)
        self.resizable(True, False)   # горизонтально тянется, вертикально — нет

        # Улучшенная поддержка DPI (HiDPI awareness)
        try:
            import ctypes
            # Пытаемся включить Per Monitor V2 (Windows 10 1703+)
            # Это предотвращает размытие при переносе между мониторами
            # -4 = DPI_AWARENESS_CONTEXT_PER_MONITOR_AWARE_V2
            ctypes.windll.user32.SetProcessDpiAwarenessContext(ctypes.c_void_p(-4))
        except Exception:
            try:
                # Фолбэк для Windows 8.1/10 (более старых билдов)
                # 2 = PROCESS_PER_MONITOR_DPI_AWARE
                ctypes.windll.shcore.SetProcessDpiAwareness(2)
            except Exception:
                try:
                    # Фолбэк для Windows Vista/7
                    ctypes.windll.user32.SetProcessDPIAware()
                except Exception:
                    pass

        # Загружаем сохранённые настройки (settings.json)
        self._settings = load_settings()

        # ─── Язык ───
        # "auto" в настройках или невалидный код → определяем из системы
        saved_lang = self._settings.get("language", "auto")
        if saved_lang == "auto" or saved_lang not in STRINGS:
            self._lang = detect_system_language()
            self._lang_auto = True
        else:
            self._lang = saved_lang
            self._lang_auto = False

        # ─── Тема ───
        saved_theme = self._settings.get("theme", "auto")
        if saved_theme == "auto":
            self._theme_name = detect_system_theme()
            self._theme_auto = True
        else:
            self._theme_name = saved_theme
            self._theme_auto = False
        
        # ─── Стиль таблицы ───
        # Загружаем сохранённый стиль или используем "s_none" по умолчанию
        self._table_style_key = self._settings.get("table_style_key", "s_none")
        
        # ─── Высота строк ───
        self._row_height_val = self._settings.get("row_height", "Авто")
        
        # ─── Состояние UI (не сохраняется) ───
        self._delim_val = self._settings.get("csv_delimiter", ";") # <-- Новое
        self._mode = "to_xlsx"
        self._csv_paths = []      # Список путей к выбранным CSV файлам
        self._pb_anim = False     # анимация прогресс-бара активна?
        self._pb_w    = 1         # текущая ширина прогресс-бара
        self._pb_pos  = 0         # положение бегунка анимации
        self._dz_text = ""        # текст в drop-zone (выставляется в _auto_load)
        self._dz_hl   = False     # подсветка drop-zone при drag-over

        # Сборка UI: тема → виджеты → автозагрузка файлов → центровка
        self._apply_theme(init=True)
        self._build()
        self._auto_load()
        self._center()

        # Перерисовка drop-zone при изменении размера окна
        self.bind("<Configure>", lambda e: (
            self._redraw_dz() if e.widget is self else None))

        # Тема titlebar применяется через after — DWM нужно реальное окно,
        # а не только tk-виджет. 50мс хватает чтобы окно появилось на экране.
        self.after(50, self._apply_title_theme)

    # ═════════════════════════════════════════════════════════════════════════
    # Тема и локализация
    # ═════════════════════════════════════════════════════════════════════════

    def _t(self) -> dict:
        """Текущая палитра (light/dark) — короткий доступ к THEMES."""
        return THEMES[self._theme_name]

    def s(self, key: str) -> str:
        """
        Локализованная строка. Если ключа нет в текущем языке — фолбэк на
        английский, если и там нет — возвращаем сам ключ (виден в UI как
        предупреждение разработчику что строку забыли перевести).
        """
        return STRINGS.get(self._lang, STRINGS["en"]).get(
            key, STRINGS["en"].get(key, key))

    def _apply_title_theme(self, target_window=None):
        """
        Переключает цвет titlebar окна (Windows 10 build 17763+).
        Использует недокументированный DWM-атрибут 20
        (DWMWA_USE_IMMERSIVE_DARK_MODE). На старых системах молча игнорируется.
        """
        win = target_window if target_window else self
        try:
            import ctypes
            # Самый надёжный способ получить HWND системной рамки окна в Tkinter
            hwnd = ctypes.windll.user32.GetParent(win.winfo_id())
            if hwnd == 0:
                hwnd = win.winfo_id()
                
            value = ctypes.c_int(1 if self._theme_name == "dark" else 0)
            ctypes.windll.dwmapi.DwmSetWindowAttribute(
                hwnd, 20, ctypes.byref(value), ctypes.sizeof(value))
            
            # Триггерим перерисовку — Windows иногда не обновляет titlebar сразу
            win.title(win.title())
        except Exception:
            pass

    def _apply_theme(self, init=False):
        """
        Применяет тему к окну. При init=True (первый вызов) перекрашивать
        ещё нечего — виджеты создаются позже в _build().
        """
        t = self._t()
        self.configure(bg=t["BG"])
        if not init:
            self._retheme_all()
            self._redraw_dz()
        self._apply_title_theme()

    def _retheme_all(self):
        """
        Перекрашивает все виджеты под текущую тему.
        Идёт рекурсивно по дереву виджетов, для каждого класса свой подход:
        Frame и Label получают фон карточки/окна в зависимости от родителя,
        Entry — фон поля ввода, Button сохраняет заданные акцентные цвета.
        """
        t = self._t()
        def walk(w):
            cls = w.winfo_class()
            try:
                if cls in ("Frame", "Labelframe"):
                    w.configure(bg=t["BG"] if w.master is self else t["CARD"])
                elif cls == "Label":
                    bg = t["BG"] if isinstance(w.master, (tk.Tk,)) else t["CARD"]
                    try:
                        cur = w.cget("bg")
                        if cur in ("#F0F4FA","#1E1E2E"):  # ранее имел цвет фона окна
                            w.configure(bg=t["BG"])
                        else:
                            w.configure(bg=t["CARD"], fg=t["TEXT"])
                    except Exception:
                        pass
                elif cls == "Entry":
                    w.configure(bg=t["ENTRY"], fg=t["TEXT"],
                                insertbackground=t["TEXT"])
                elif cls == "Button":
                    cur_bg = w.cget("bg")
                    # Keep coloured buttons (green, accent2) as-is
                    if cur_bg not in ("#4A7C59","#2E5FA3","#3A6FBF","#1F3864","#0F1F40"):
                        w.configure(bg=t["CARD"], fg=t["ACC2"])
                elif cls == "Canvas":
                    if w is not self.drop_canvas and w is not self.pb_canvas:
                        w.configure(bg=t["CARD"])
            except Exception:
                pass
            for child in w.winfo_children():
                walk(child)
        walk(self)
        
        # Special widgets — применяем явно для надёжности
        self.configure(bg=t["BG"])
        self.hdr_frame.configure(bg=t["ACCENT"])
        self.hdr_top_frame.configure(bg=t["ACCENT"])
        HDR_FG = "#C8D8F0"
        for w in self.hdr_labels:
            try:
                cls = w.winfo_class()
                w.configure(bg=t["ACCENT"])
                if cls == "Button":
                    w.configure(fg=HDR_FG, activebackground="#3A5A8A")
                elif cls == "Label":
                    w.configure(fg="white")
            except Exception:
                pass
        self.card.configure(bg=t["CARD"],
                            highlightbackground=t["BORDER"])
        # Перекрасить все дочерние Frame и Label внутри card
        for child in self.card.winfo_children():
            try:
                cls = child.winfo_class()
                if cls == "Frame":
                    child.configure(bg=t["CARD"])
                    for gc in child.winfo_children():
                        try:
                            gcls = gc.winfo_class()
                            if gcls == "Label":
                                gc.configure(bg=t["CARD"], fg=t["TEXT"])
                            elif gcls == "Button":
                                gbg = gc.cget("bg")
                                if gbg not in ("#4A7C59", "#2E5E3E"):
                                    gc.configure(bg=t["CARD"], fg=t["ACC2"])
                            elif gcls == "Frame":
                                gc.configure(bg=t["CARD"])
                        except Exception:
                            pass
                elif cls == "Label":
                    child.configure(bg=t["CARD"], fg=t["TEXT"])
                elif cls == "Button":
                    cbg = child.cget("bg")
                    if cbg not in ("#4A7C59", "#2E5E3E", t["ACC2"]):
                        child.configure(bg=t["CARD"], fg=t["ACC2"])
                elif cls == "Entry":
                    child.configure(bg=t["ENTRY"], fg=t["TEXT"])
            except Exception:
                pass
        self.status_frame.configure(bg=t["BG"])
        self.status_lbl.configure(bg=t["BG"], fg=t["MUTED"])
        self.pb_frame.configure(bg=t["BG"])
        self.pb_canvas.configure(bg=t["PB_BG"])
        self.pb_canvas.itemconfig(self.pb_rect, fill=t["ACC2"])
        self.hint_lbl.configure(bg=t["CARD"])
        # Кастомное поле выбора стиля
        self.style_field.configure(bg=t["ENTRY"])
        self.style_lbl_val.configure(bg=t["ENTRY"], fg=t["TEXT"])
        self.style_arrow.configure(bg=t["ENTRY"], fg=t["MUTED"])
        # Кастомные поля (стиль и высота строк/разделитель)
        if hasattr(self, 'settings_row'):
            self.settings_row.configure(bg=t["CARD"])
            self.rh_frame.configure(bg=t["ENTRY"])
            self.rh_entry.configure(bg=t["ENTRY"], fg=t["TEXT"], insertbackground=t["TEXT"])
            self.rh_arrow.configure(bg=t["ENTRY"], fg=t["MUTED"])
            self.dlm_frame.configure(bg=t["ENTRY"])
            self.dlm_entry.configure(bg=t["ENTRY"], fg=t["TEXT"], insertbackground=t["TEXT"])
            self.dlm_arrow.configure(bg=t["ENTRY"], fg=t["MUTED"])
        self._refresh_hint()

    def _toggle_theme(self):
        """Переключает тему по клику на 🌙/☀, сохраняет выбор в settings.json."""
        self._theme_name = "dark" if self._theme_name == "light" else "light"
        self._theme_auto = False   # явный выбор пользователя — больше не auto
        self._settings["theme"] = self._theme_name
        save_settings(self._settings)
        self._apply_theme()
        self._update_toggle_btn()

    def _update_toggle_btn(self):
        """Обновляет иконку кнопки темы (☀ когда темно, 🌙 когда светло)."""
        t = self._t()
        if self._theme_name == "dark":
            self.theme_btn.config(text="☀", fg="#FFD700", bg=t["ACCENT"],
                                  activebackground="#3A5A8A")
        else:
            self.theme_btn.config(text="🌙", fg="#C8D8F0", bg=t["ACCENT"],
                                  activebackground="#3A5A8A")

    # ═════════════════════════════════════════════════════════════════════════
    # Диалоги: О программе, Выбор языка
    # ═════════════════════════════════════════════════════════════════════════

    def _show_about(self):
        """Модальное окно с описанием программы и кликабельной TG-ссылкой."""
        t = self._t()
        win = tk.Toplevel(self)
        win.title(self.s("about_ttl"))
        win.configure(bg=t["CARD"])
        win.resizable(False, False)
        win.transient(self)   # связан с главным окном (закроется вместе с ним)
        win.grab_set()        # модальный — блокирует главное окно

        tk.Label(win, text=self.s("title"),
                 font=(FONT, 13, "bold"), fg=t["TEXT"], bg=t["CARD"]
                 ).pack(pady=(20, 4), padx=30)
        tk.Label(win, text=self.s("about_ver").format(ver=VERSION),
                 font=(FONT, 9), fg=t["MUTED"], bg=t["CARD"]
                 ).pack(pady=(0, 10))

        tk.Frame(win, bg=t["BORDER"], height=1).pack(fill="x", padx=20)

        tk.Label(win, text=self.s("about_desc"),
                 font=(FONT, 9), fg=t["TEXT"], bg=t["CARD"],
                 justify="left", anchor="w", padx=24, pady=12
                 ).pack(fill="x")

        tk.Frame(win, bg=t["BORDER"], height=1).pack(fill="x", padx=20)

        # Telegram link
        tg_frame = tk.Frame(win, bg=t["CARD"])
        tg_frame.pack(fill="x", padx=24, pady=(10, 4))
        tk.Label(tg_frame, text="✉  ", font=(FONT, 9),
                 fg=t["MUTED"], bg=t["CARD"]).pack(side="left")
        tg_lbl = tk.Label(tg_frame, text="t.me/maxsteff",
                           font=(FONT, 9, "underline"),
                           fg="#2E5FA3", bg=t["CARD"], cursor="hand2")
        tg_lbl.pack(side="left")
        tg_lbl.bind("<Button-1>", lambda _: os.startfile("https://t.me/maxsteff"))

        tk.Frame(win, bg=t["BORDER"], height=1).pack(fill="x", padx=20, pady=(10, 0))
        tk.Button(win, text=self.s("close"),
                  font=(FONT, 9), fg="white", bg=t["ACC2"],
                  activebackground="#1A4A8A", activeforeground="white",
                  relief="flat", bd=0, padx=20, pady=6,
                  cursor="hand2", command=win.destroy).pack(pady=14)

        win.update_idletasks()
        pw, ph = self.winfo_x(), self.winfo_y()
        ww = self.winfo_width()
        win.geometry(f"+{pw + ww//2 - win.winfo_reqwidth()//2}+{ph + 60}")
        
        # Добавляем покраску заголовка всплывающего окна
        self.after(50, lambda: self._apply_title_theme(win))

    def _show_lang(self):
        """Диалог выбора языка с радиокнопками."""
        t = self._t()
        win = tk.Toplevel(self)
        win.title(self.s("lang_ttl"))
        win.configure(bg=t["CARD"])
        win.resizable(False, False)
        win.transient(self)
        win.grab_set()

        tk.Label(win, text=self.s("lang_ttl"),
                 font=(FONT, 11, "bold"), fg=t["TEXT"], bg=t["CARD"]
                 ).pack(pady=(16, 8), padx=24)

        # Радиокнопки. Текущий язык помечается "(определён автоматически)"
        # если был выбран автоматически, чтобы пользователь видел почему именно он
        lang_var = tk.StringVar(value=self._lang)
        frm = tk.Frame(win, bg=t["CARD"])
        frm.pack(fill="x", padx=24, pady=4)
        for code, name in LANGUAGES.items():
            auto_mark = (f"  {self.s('lang_auto')}"
                         if self._lang_auto and code == self._lang else "")
            tk.Radiobutton(frm, text=f"{name}{auto_mark}",
                           variable=lang_var, value=code,
                           font=(FONT, 9), fg=t["TEXT"], bg=t["CARD"],
                           selectcolor=t["CARD"], activebackground=t["CARD"],
                           cursor="hand2").pack(anchor="w", pady=2)

        tk.Frame(win, bg=t["BORDER"], height=1).pack(fill="x", padx=16, pady=8)

        def apply_lang():
            """Сохраняет выбор и пересобирает UI на новом языке."""
            chosen = lang_var.get()
            self._lang = chosen
            self._lang_auto = False
            self._settings["language"] = chosen
            save_settings(self._settings)
            win.destroy()
            self._apply_lang()

        tk.Button(win, text=self.s("apply"),
                  font=(FONT, 9), fg="white", bg=t["ACC2"],
                  activebackground="#1A4A8A", activeforeground="white",
                  relief="flat", bd=0, padx=20, pady=6,
                  cursor="hand2", command=apply_lang).pack(pady=(0, 14))

        # Позиционируем диалог по центру главного окна, ниже шапки
        win.update_idletasks()
        pw, ph = self.winfo_x(), self.winfo_y()
        ww = self.winfo_width()
        win.geometry(f"+{pw + ww//2 - win.winfo_reqwidth()//2}+{ph + 60}")
        # Добавляем покраску заголовка всплывающего окна
        self.after(50, lambda: self._apply_title_theme(win))
        
    def _apply_lang(self):
        """
        Обновляет все тексты в UI без пересоздания виджетов.
        Перебирает все элементы у которых текст зависит от языка.
        """
        # Заголовок окна (titlebar) и большой заголовок в шапке
        self.title(self.s("title"))
        self.hdr_labels[0].config(text=self.s("title"))

        # Кнопки шапки
        self.about_btn.config(text=self.s("about_btn"))
        self.lang_btn_w.config(text=self.s("lang_btn"))

        # Метки полей в карточке
        self._update_card_labels()

        # Статус-строка
        self.status_var.set(self.s("status_idle"))
        self.status_lbl.config(fg=self._t()["MUTED"])

        # Текст в drop-zone
        self._dz_text = self.s("dz_idle")
        self._redraw_dz()

        # Селектор стиля таблицы — список и текущая метка
        new_names = get_table_style_names(self._lang)
        new_label = STRINGS.get(self._lang, STRINGS["en"]).get(
            self._table_style_key, new_names[0])
        self.style_var.set(new_label)

    def _update_card_labels(self):
        """Обновляет текст меток в карточке в зависимости от языка и режима."""
        keys = ["csv_lbl", "xlsx_lbl", "col_lbl", "style_lbl", "rh_only_lbl"]
        
        # Меняем ключи строк, если включен режим обратной конвертации
        if getattr(self, '_mode', 'to_xlsx') == 'to_csv':
            keys[0] = "xlsx_in_lbl"
            keys[1] = "csv_out_lbl"
            keys[4] = "dlm_only_lbl"
            
        rows = [0, 2, 4, 7, 9]   # Строки, на которых лежат эти подписи
        for key, row in zip(keys, rows):
            for child in self.card.grid_slaves(row=row, column=0):
                if child.winfo_class() == "Label":
                    child.config(text=self.s(key))

        # Кнопки внутри карточки
        self.btn.config(text=self.s("convert_btn"))
        self.create_btn.config(text=self.s("create_ex"))

        self._refresh_hint()
        self._update_visibility()  # <--- Запускаем умное скрытие
        
    def _update_visibility(self):
        """Скрывает или показывает элементы интерфейса в зависимости от режима."""
        is_to_csv = getattr(self, '_mode', 'to_xlsx') == 'to_csv'
        
        # Помощник для получения Label заголовка по номеру строки
        def get_lbl(row):
            slaves = self.card.grid_slaves(row=row, column=0)
            for s in slaves:
                if s.winfo_class() == "Label": return s
            return None

        col_lbl = get_lbl(4)
        style_lbl = get_lbl(7)

        if is_to_csv:
            # === РЕЖИМ .XLSX -> .CSV ===
            # Скрываем настройки Excel (Фильтр, Стиль, Высоту строк)
            if col_lbl: col_lbl.grid_remove()
            if hasattr(self, 'col_entry'): self.col_entry.grid_remove()
            if hasattr(self, 'col_btn_frame'): self.col_btn_frame.grid_remove()
            if hasattr(self, 'hint_lbl'): self.hint_lbl.grid_remove()
            
            if style_lbl: style_lbl.grid_remove()
            if hasattr(self, 'style_frame'): self.style_frame.grid_remove()
            
            if hasattr(self, 'rh_frame'): self.rh_frame.pack_forget()
            # Показываем только разделитель
            if hasattr(self, 'dlm_frame'): self.dlm_frame.pack(side="left")
        else:
            # === РЕЖИМ .CSV -> .XLSX ===
            # Показываем настройки Excel
            if col_lbl: col_lbl.grid()
            if hasattr(self, 'col_entry'): self.col_entry.grid()
            if hasattr(self, 'col_btn_frame'): self.col_btn_frame.grid()
            if hasattr(self, 'hint_lbl'): self.hint_lbl.grid()
            
            if style_lbl: style_lbl.grid()
            if hasattr(self, 'style_frame'): self.style_frame.grid()
            
            # Скрываем разделитель, показываем высоту строк
            if hasattr(self, 'dlm_frame'): self.dlm_frame.pack_forget()
            if hasattr(self, 'rh_frame'): self.rh_frame.pack(side="left", padx=(0, 20))
            
        # Магия: принудительно подгоняем размер окна по высоте, 
        # чтобы оно элегантно сжалось после скрытия элементов
        def adjust():
            self.update_idletasks()
            self.geometry(f"{self.winfo_width()}x{self.winfo_reqheight()}")
        self.after(10, adjust)

    # ═════════════════════════════════════════════════════════════════════════
    # Построение UI
    # ═════════════════════════════════════════════════════════════════════════

    def _build(self):
        """
        Строит всё дерево виджетов. Вызывается один раз из __init__.
        Все виджеты которые потом нужно ретемировать или релокализовать
        сохраняются в self.* для последующего доступа.
        """
        t = self._t()

        # ── Header (синяя шапка) ──────────────────────────────────────────
        self.hdr_frame = tk.Frame(self, bg=t["ACCENT"])
        self.hdr_frame.pack(fill="x")

        # Header top row: about | lang | ... | theme-toggle
        self.hdr_top_frame = tk.Frame(self.hdr_frame, bg=t["ACCENT"])
        self.hdr_top_frame.pack(fill="x", padx=14, pady=(10, 0))

        HDR_FG = "#C8D8F0"   # светлый цвет для кнопок в шапке
        HDR_ACTIVE = "#3A5A8A"

        # About button
        self.about_btn = tk.Button(
            self.hdr_top_frame, text=self.s("about_btn"),
            font=(FONT, 8), fg=HDR_FG,
            bg=t["ACCENT"], activebackground=HDR_ACTIVE,
            activeforeground="white",
            relief="flat", bd=0, padx=6, pady=2, cursor="hand2",
            command=self._show_about)
        self.about_btn.pack(side="left", padx=(0, 6))

        # Language button
        self.lang_btn_w = tk.Button(
            self.hdr_top_frame, text=self.s("lang_btn"),
            font=(FONT, 8), fg=HDR_FG,
            bg=t["ACCENT"], activebackground=HDR_ACTIVE,
            activeforeground="white",
            relief="flat", bd=0, padx=6, pady=2, cursor="hand2",
            command=self._show_lang)
        self.lang_btn_w.pack(side="left")

        # Theme toggle (right side)
        self.theme_btn = tk.Button(
            self.hdr_top_frame, text="", font=(FONT, 13),
            bg=t["ACCENT"], activebackground=HDR_ACTIVE,
            relief="flat", bd=0, padx=4, cursor="hand2",
            command=self._toggle_theme)
        self.theme_btn.pack(side="right")
        self._update_toggle_btn()

        # Title labels
        lbl1 = tk.Label(self.hdr_frame, text=self.s("title"),
                        font=(FONT, 15, "bold"), fg="white",
                        bg=t["ACCENT"], pady=4)
        lbl1.pack()
        lbl2 = tk.Label(self.hdr_frame, text=self.s("subtitle"),
                        font=(FONT, 9), fg="#A8BFDF",
                        bg=t["ACCENT"])
        lbl2.pack(pady=(0, 10))
        self.hdr_labels = [lbl1, lbl2, self.about_btn,
                           self.lang_btn_w, self.theme_btn]

        # Drop zone
        dz_wrap = tk.Frame(self, bg=t["BG"])
        dz_wrap.pack(fill="x", padx=20, pady=(16, 0))
        self.drop_canvas = tk.Canvas(dz_wrap, height=80,
                                     bg=t["DZ_BG"], highlightthickness=0,
                                     cursor="hand2")
        self.drop_canvas.pack(fill="x")
        self.drop_canvas.bind("<Button-1>",  lambda _: self._pick_csv())
        self.drop_canvas.bind("<Configure>", lambda e: self._redraw_dz())
        if _DND:
            self.drop_canvas.drop_target_register(DND_FILES)
            self.drop_canvas.dnd_bind("<<Drop>>",      self._on_drop)
            self.drop_canvas.dnd_bind("<<DragEnter>>", self._on_drag_enter)
            self.drop_canvas.dnd_bind("<<DragLeave>>", self._on_drag_leave)

        # Card
        self.card = tk.Frame(self, bg=t["CARD"], padx=24, pady=20,
                             highlightthickness=1,
                             highlightbackground=t["BORDER"])
        self.card.pack(fill="x", padx=20, pady=14)
        self.card.columnconfigure(0, weight=1)

        # CSV
        self._lbl("CSV-файл (экспорт Jira):", 0)
        self.csv_var = tk.StringVar()
        self._row(self.csv_var, 1, self._pick_csv, pady_b=12)

        # XLSX
        self._lbl("Сохранить XLSX как:", 2)
        self.xlsx_var = tk.StringVar()
        self._row(self.xlsx_var, 3, self._pick_xlsx, pady_b=12)

        # Columns
        # Columns
        self._lbl("Фильтр колонок (columns.txt):", 4)
        self.col_var = tk.StringVar()
        self.col_entry = tk.Entry(self.card, textvariable=self.col_var,
                                  font=(FONT, 9), relief="solid", bd=1,
                                  bg=t["ENTRY"], fg=t["TEXT"])
        self.col_entry.grid(row=5, column=0, sticky="ew", padx=(0, 8), pady=(0, 4), ipady=4)
        
        self.col_btn_frame = tk.Frame(self.card, bg=t["CARD"])
        self.col_btn_frame.grid(row=5, column=1, pady=(0, 4), sticky="e")
        tk.Button(self.col_btn_frame, text=self.s("browse"), font=(FONT, 9), fg=t["ACC2"],
                  bg=t["CARD"], relief="solid", bd=1, padx=10, cursor="hand2",
                  command=self._pick_columns).pack(side="left", padx=(0, 6))
        self.create_btn = tk.Button(self.col_btn_frame, text=self.s("create_ex"),
                                    font=(FONT, 9), fg="white", bg="#4A7C59",
                                    activebackground="#2E5E3E", activeforeground="white",
                                    relief="flat", bd=0, padx=10, pady=2, cursor="hand2",
                                    command=self._create_example)
        self.create_btn.pack(side="left")
        self._hover(self.create_btn, "#4A7C59", "#2E5E3E")

        # Hint
        self.hint_var = tk.StringVar()
        self.hint_lbl = tk.Label(self.card, textvariable=self.hint_var,
                                 font=(FONT, 8), fg=t["MUTED"], bg=t["CARD"],
                                 anchor="w")
        self.hint_lbl.grid(row=6, column=0, columnspan=2,
                           sticky="w", pady=(0, 8))

        # Table style selector — кастомный picker с миниатюрами
        self._lbl(self.s("style_lbl"), 7)
        self.style_frame = tk.Frame(self.card, bg=t["CARD"])
        self.style_frame.grid(row=8, column=0, columnspan=2, sticky="ew", pady=(0, 12))
        self.style_frame.columnconfigure(0, weight=1)

        _names = get_table_style_names(self._lang)
        _cur_label = STRINGS.get(self._lang, STRINGS["en"]).get(
            self._table_style_key, _names[0])
        self.style_var = tk.StringVar(value=_cur_label)

        # Кликабельное поле (имитация combobox)
        self.style_field = tk.Frame(self.style_frame, bg=t["ENTRY"],
                                    relief="solid", bd=1, cursor="hand2")
        self.style_field.grid(row=0, column=0, sticky="ew")
        self.style_field.columnconfigure(0, weight=1)
        self.style_lbl_val = tk.Label(self.style_field,
                                      textvariable=self.style_var,
                                      font=(FONT, 9), fg=t["TEXT"],
                                      bg=t["ENTRY"], anchor="w",
                                      padx=6, pady=5)
        self.style_lbl_val.grid(row=0, column=0, sticky="ew")
        self.style_arrow = tk.Label(self.style_field, text="▼",
                                    font=(FONT, 7), fg=t["MUTED"],
                                    bg=t["ENTRY"], padx=6)
        self.style_arrow.grid(row=0, column=1)
        self.style_arrow.grid(row=0, column=1)
        for w in (self.style_field, self.style_lbl_val, self.style_arrow):
            w.bind("<Button-1>", lambda _: self._show_style_picker())

        # ─── Настройки: Высота строк + Разделитель (в один ряд) ───
        self._lbl(self.s("rh_only_lbl"), 9)
        
        self.settings_row = tk.Frame(self.card, bg=t["CARD"])
        self.settings_row.grid(row=10, column=0, columnspan=2, sticky="w", pady=(0, 12))

        # 1. Высота строк
        self.rh_var = tk.StringVar(value=self._row_height_val)
        self.rh_frame = tk.Frame(self.settings_row, bg=t["ENTRY"], relief="solid", bd=1)
        self.rh_frame.pack(side="left", padx=(0, 20))
        self.rh_entry = tk.Entry(self.rh_frame, textvariable=self.rh_var, font=(FONT, 9), relief="flat", bd=0, bg=t["ENTRY"], fg=t["TEXT"], width=10)
        self.rh_entry.pack(side="left", padx=(6, 0), pady=4)
        self.rh_arrow = tk.Label(self.rh_frame, text="▼", font=(FONT, 7), fg=t["MUTED"], bg=t["ENTRY"], cursor="hand2")
        self.rh_arrow.pack(side="right", padx=6)

        # 2. Разделитель CSV
        self.dlm_var = tk.StringVar(value=self._delim_val)
        self.dlm_frame = tk.Frame(self.settings_row, bg=t["ENTRY"], relief="solid", bd=1)
        self.dlm_frame.pack(side="left")
        self.dlm_entry = tk.Entry(self.dlm_frame, textvariable=self.dlm_var, font=(FONT, 9), relief="flat", bd=0, bg=t["ENTRY"], fg=t["TEXT"], width=6)
        self.dlm_entry.pack(side="left", padx=(6, 0), pady=4)
        self.dlm_arrow = tk.Label(self.dlm_frame, text="▼", font=(FONT, 7), fg=t["MUTED"], bg=t["ENTRY"], cursor="hand2")
        self.dlm_arrow.pack(side="right", padx=6)

        # Обработчики выпадающих меню
        def create_popup(frame_widget, var_widget, options, setting_key):
            if hasattr(self, '_active_popup') and self._active_popup.winfo_exists():
                self._active_popup.destroy()
            popup = tk.Toplevel(self)
            popup.overrideredirect(True)
            popup.configure(bg=self._t()["CARD"], highlightthickness=1, highlightbackground=self._t()["BORDER"])
            self._active_popup = popup
            
            x, y, w = frame_widget.winfo_rootx(), frame_widget.winfo_rooty() + frame_widget.winfo_height(), frame_widget.winfo_width()
            popup.geometry(f"+{x}+{y}")
            popup.minsize(w, 1)

            for val in options:
                lbl = tk.Label(popup, text=val, font=(FONT, 9), bg=self._t()["CARD"], fg=self._t()["TEXT"], anchor="w", padx=10, pady=4, cursor="hand2")
                lbl.pack(fill="x")
                lbl.bind("<Enter>", lambda e, l=lbl: l.config(bg=self._t()["ACC2"], fg="white"))
                lbl.bind("<Leave>", lambda e, l=lbl: l.config(bg=self._t()["CARD"], fg=self._t()["TEXT"]))
                def on_click(e, v=val):
                    var_widget.set(v)
                    self._settings[setting_key] = v
                    save_settings(self._settings)
                    popup.destroy()
                lbl.bind("<Button-1>", on_click)

            popup.bind("<Button-1>", lambda e: popup.destroy() if not (popup.winfo_rootx() <= e.x_root <= popup.winfo_rootx() + popup.winfo_width() and popup.winfo_rooty() <= e.y_root <= popup.winfo_rooty() + popup.winfo_height()) else None)
            popup.bind("<FocusOut>", lambda e: popup.destroy() if e.widget == popup else None)
            popup.focus_set()
            popup.grab_set()

        self.rh_arrow.bind("<Button-1>", lambda e: create_popup(self.rh_frame, self.rh_var, ["Авто", "15", "20", "30", "45", "60"], "row_height"))
        self.dlm_arrow.bind("<Button-1>", lambda e: create_popup(self.dlm_frame, self.dlm_var, [";", ",", "|", "Tab"], "csv_delimiter"))

        def save_typing(*args):
            self._settings["row_height"], self._settings["csv_delimiter"] = self.rh_var.get(), self.dlm_var.get()
            save_settings(self._settings)
            
        self.rh_entry.bind("<KeyRelease>", save_typing)
        self.dlm_entry.bind("<KeyRelease>", save_typing)

        # Divider (сдвигаем на 11 строку)
        tk.Frame(self.card, bg=t["BORDER"], height=1).grid(
            row=11, column=0, columnspan=2, sticky="ew", pady=(0, 14))

        # Convert button (сдвигаем на 12 строку)
        self.btn = tk.Button(
            self.card, text="  ▶  Преобразовать",
            font=(FONT, 11, "bold"), fg="white", bg=t["ACC2"],
            activebackground="#1A4A8A", activeforeground="white",
            relief="flat", bd=0, padx=36, pady=11,
            cursor="hand2", command=self._run)
        self.btn.grid(row=12, column=0, columnspan=2, pady=(0, 4))
        self._hover(self.btn, t["ACC2"], "#1A4A8A")

        # Status
        self.status_frame = tk.Frame(self, bg=t["BG"])
        self.status_frame.pack(fill="x", padx=20, pady=(6, 2))
        self.status_var = tk.StringVar(value="Выберите файлы и нажмите «Преобразовать»")
        self.status_lbl = tk.Label(self.status_frame, textvariable=self.status_var,
                                   font=(FONT, 9), fg=t["MUTED"],
                                   bg=t["BG"], anchor="w")
        self.status_lbl.pack(fill="x")

        # Progress bar
        self.pb_frame = tk.Frame(self, bg=t["BG"])
        self.pb_frame.pack(fill="x", padx=20, pady=(4, 16))
        self.pb_canvas = tk.Canvas(self.pb_frame, height=6, bg=t["PB_BG"],
                                   highlightthickness=0)
        self.pb_canvas.pack(fill="x")
        self.pb_canvas.bind("<Configure>",
                            lambda e: setattr(self, "_pb_w", max(e.width, 1)))
        self.pb_rect = self.pb_canvas.create_rectangle(
            0, 0, 0, 6, fill=t["ACC2"], outline="")

        self._refresh_hint()
        self._update_visibility()

    # ── Helpers ───────────────────────────────────────────────────────────────

    # ─── Хелперы построения UI ───────────────────────────────────────────

    def _lbl(self, text, row):
        """Жирная подпись над полем ввода в карточке."""
        t = self._t()
        tk.Label(self.card, text=text, font=(FONT, 9, "bold"),
                 fg=t["TEXT"], bg=t["CARD"], anchor="w").grid(
            row=row, column=0, columnspan=2, sticky="w", pady=(0, 4))

    def _row(self, var, row, cmd, pady_b=14):
        """
        Стандартная строка "поле ввода + кнопка Обзор".
        Кнопка обёрнута в Frame чтобы grid-колонки полей CSV/XLSX/columns
        имели одинаковую ширину независимо от наличия "Создать пример".
        """
        t = self._t()
        tk.Entry(self.card, textvariable=var, font=(FONT, 9),
                 relief="solid", bd=1, bg=t["ENTRY"], fg=t["TEXT"]).grid(
            row=row, column=0, sticky="ew",
            padx=(0, 8), pady=(0, pady_b), ipady=4)
        bf = tk.Frame(self.card, bg=t["CARD"])
        bf.grid(row=row, column=1, pady=(0, pady_b), sticky="ew")
        tk.Button(bf, text="Обзор…", font=(FONT, 9), fg=t["ACC2"],
                  bg=t["CARD"], relief="solid", bd=1, padx=10, cursor="hand2",
                  command=cmd).pack(side="left")

    def _hover(self, w, n, h):
        """Привязывает hover-эффект к виджету: nrm→hover на наведении."""
        w.bind("<Enter>", lambda _: w.config(bg=h))
        w.bind("<Leave>", lambda _: w.config(bg=n))

    def _center(self):
        """Центрирует окно на экране и устанавливает минимальный размер."""
        self.update_idletasks()
        w = max(self.winfo_reqwidth(), MIN_W)
        h = self.winfo_reqheight()
        self.geometry(f"{w}x{h}+"
                      f"{(self.winfo_screenwidth()-w)//2}+"
                      f"{(self.winfo_screenheight()-h)//2}")
        self.minsize(MIN_W, h)

    # ═════════════════════════════════════════════════════════════════════════
    # Drop zone (область для перетаскивания CSV)
    # ═════════════════════════════════════════════════════════════════════════

    def _redraw_dz(self):
        """
        Перерисовывает drop-zone. Пунктирная рамка строится из 4 line'ов
        вместо create_rectangle с dash — у Canvas нет родного pattern для
        прямоугольных пунктиров, line даёт единообразный результат.
        Состояние подсветки (drag-over) хранится в self._dz_hl.
        """
        t  = self._t()
        c  = self.drop_canvas
        cw = c.winfo_width()
        ch = c.winfo_height()
        if cw < 2:
            return  # canvas ещё не получил размеры
        c.delete("all")
        hl  = self._dz_hl
        bg  = "#D6E8FF" if hl else t["DZ_BG"]
        clr = "#3A7BD5" if hl else t["DZ_CL"]
        c.configure(bg=bg)
        # 4 пунктирные линии вместо одного create_rectangle с dash
        pad = 5
        kw = dict(fill=clr, dash=(6, 4), width=2)
        c.create_line(pad,    pad,    cw-pad, pad,    **kw)   # top
        c.create_line(cw-pad, pad,    cw-pad, ch-pad, **kw)   # right
        c.create_line(cw-pad, ch-pad, pad,    ch-pad, **kw)   # bottom
        c.create_line(pad,    ch-pad, pad,    pad,    **kw)   # left
        # Иконка + подпись по центру
        icon = "📂" if hl else "📄"
        c.create_text(cw//2, ch//2 - 10, text=icon,
                      font=(FONT, 18), fill=clr)
        c.create_text(cw//2, ch//2 + 14, text=self._dz_text,
                      font=(FONT, 9),
                      fill="#1A4A8A" if hl else t["DZ_TXT"])

    def _on_drop(self, event):
        """Обработчик drop — приходит из tkinterdnd2."""
        # tk.splitlist идеально парсит Tcl-список путей (даже с пробелами)
        paths = self.tk.splitlist(event.data)
        self._load_csv(paths)

    def _on_drag_enter(self, _):
        """Файл "вошёл" в drop-zone — меняем подсказку и подсвечиваем."""
        self._dz_text = self.s("dz_drop")
        self._dz_hl   = True
        self._redraw_dz()

    def _on_drag_leave(self, _):
        """Файл ушёл — возвращаем предыдущее состояние."""
        n = len(getattr(self, '_csv_paths', []))
        if n == 1:
            self._dz_text = f"📄  {Path(self._csv_paths[0]).name}"
        elif n > 1:
            self._dz_text = f"📄  Выбрано файлов: {n}"
        else:
            self._dz_text = self.s("dz_idle")
        self._dz_hl = False
        self._redraw_dz()

    def _load_csv(self, paths):
        if isinstance(paths, str): paths = [paths]
        valid_paths = [p for p in paths if p.lower().endswith((".csv", ".xlsx"))]
        if not valid_paths:
            messagebox.showwarning(self.s("wrong_fmt_ttl"), self.s("wrong_fmt"))
            return
            
        self._csv_paths = valid_paths
        
        # МАГИЯ: Определяем режим по расширению первого закинутого файла
        first_ext = Path(valid_paths[0]).suffix.lower()
        self._mode = "to_csv" if first_ext == ".xlsx" else "to_xlsx"
        out_ext = ".csv" if self._mode == "to_csv" else ".xlsx"
        
        self._update_card_labels() # Перерисовываем подписи
        
        if len(valid_paths) == 1:
            self.csv_var.set(valid_paths[0])
            self._dz_text = f"📄  {Path(valid_paths[0]).name}"
            
            cur_out = self.xlsx_var.get()
            if not cur_out or Path(cur_out).suffix.lower() != out_ext:
                self.xlsx_var.set(str(Path(valid_paths[0]).with_suffix(out_ext)))
        else:
            loc_map = {"ru": "Выбрано файлов", "en": "Selected files"}
            msg = f"{loc_map.get(self._lang, 'Selected files')}: {len(valid_paths)}"
            self.csv_var.set(msg)
            self._dz_text = f"📄  {msg}"
            
            cur_out = self.xlsx_var.get()
            if not cur_out or Path(cur_out).suffix.lower() != out_ext:
                parent = Path(valid_paths[0]).parent
                self.xlsx_var.set(str(parent / f"Jira_Merged_Export{out_ext}"))
                
        self._dz_hl = False
        self._redraw_dz()

    def _pick_csv(self):
        paths = filedialog.askopenfilenames(
            title=self.s("pick_csv_ttl"),
            filetypes=[("Data files", "*.csv *.xlsx"), (self.s("all_ft"), "*.*")])
        if paths: self._load_csv(paths)

    def _pick_xlsx(self):
        ini = self.xlsx_var.get() or os.path.expanduser("~")
        out_ext = ".csv" if getattr(self, '_mode', 'to_xlsx') == 'to_csv' else ".xlsx"
        ft_name = self.s("csv_ft") if out_ext == ".csv" else self.s("xlsx_ft")
        
        p = filedialog.asksaveasfilename(
            title=self.s("pick_xlsx_ttl"),
            initialfile=Path(ini).name if self.xlsx_var.get() else f"output{out_ext}",
            initialdir=str(Path(ini).parent),
            defaultextension=out_ext,
            filetypes=[(ft_name, f"*{out_ext}"), (self.s("all_ft"), "*.*")])
        if p: self.xlsx_var.set(p)

    def _pick_columns(self):
        p = filedialog.askopenfilename(
            title=self.s("pick_col_ttl"),
            filetypes=[(self.s("txt_ft"), "*.txt"), (self.s("all_ft"), "*.*")])
        if p:
            self.col_var.set(p)
            self._refresh_hint()

    def _create_example(self):
        """
        Создаёт columns.txt-шаблон рядом с .exe (или рядом с уже выбранным
        columns.txt). Берёт заголовки текущего CSV, применяет auto_rename.
        """
        # 1. Проверяем, есть ли у нас вообще выбранные файлы в списке
        if not hasattr(self, '_csv_paths') or not self._csv_paths:
            messagebox.showwarning(self.s("err_ttl"), self.s("no_csv_ex"))
            return
            
        # 2. Берем первый файл из списка для генерации шаблона колонок
        csv_path = self._csv_paths[0]
        if not Path(csv_path).exists():
            messagebox.showwarning(self.s("err_ttl"), self.s("no_csv_ex"))
            return
            
        col_path = self.col_var.get().strip()
        save_dir = Path(col_path).parent if col_path else get_exe_dir()
        out = save_dir / "columns.txt"
        
        if out.exists():
            messagebox.showinfo(self.s("exists_ttl"),
                                self.s("already_exists").format(path=out))
            return
            
        _, strips = load_columns_filter(col_path)
        dlm = self.dlm_var.get().strip()
        
        # Передаем разделитель в функцию генерации
        result = generate_columns_example(csv_path, save_dir, strips or None, delimiter=dlm)
        
        if result:
            if messagebox.askyesno(self.s("created_ttl"),
                                   self.s("created_msg").format(path=result)):
                os.startfile(str(Path(result).parent))
        else:
            messagebox.showerror(self.s("create_err_ttl"), self.s("create_err"))

    def _on_style_change(self, _=None):
        """Обработчик изменения стиля через виджет (резерв для будущего)."""
        label = self.style_var.get()
        key = get_style_key_by_label(label, self._lang)
        if key:
            self._table_style_key = key
            self._settings["table_style_key"] = key
            save_settings(self._settings)

    # ═════════════════════════════════════════════════════════════════════════
    # Селектор стилей таблицы (миниатюры)
    # ═════════════════════════════════════════════════════════════════════════

    def _draw_style_thumb(self, canvas, key, w, h, selected=False):
        """
        Рисует на Canvas мини-превью таблицы для одного стиля.
        Структура: 1 строка заголовка + 4 строки данных, 4 колонки.
        Для имитации текста — горизонтальные линии в каждой ячейке.

        Параметры:
            canvas    — целевой tk.Canvas (его размеры выставляются здесь)
            key       — ключ стиля из ALL_TABLE_STYLES
            w, h      — размеры миниатюры в пикселях
            selected  — рисовать ли акцентную рамку выделения
        """
        colors = TABLE_STYLE_COLORS.get(key)
        if not colors:
            return
        hdr_c, odd_c, even_c, txt_c = colors
        t = self._t()
        bg = t["CARD"]

        canvas.delete("all")
        canvas.configure(bg=bg, width=w, height=h)

        ROWS = 5          # 1 заголовок + 4 строки данных
        COLS = 4
        PAD  = 2

        cell_h = max(1, (h - PAD * 2) // ROWS)
        cell_w = max(1, (w - PAD * 2) // COLS)

        for ri in range(ROWS):
            y0 = PAD + ri * cell_h
            y1 = y0 + cell_h
            # Чередование: 0=заголовок, нечётные=odd, чётные=even
            fill = hdr_c if ri == 0 else (odd_c if ri % 2 == 1 else even_c)

            for ci in range(COLS):
                x0 = PAD + ci * cell_w
                x1 = x0 + cell_w

                canvas.create_rectangle(x0, y0, x1, y1,
                                        fill=fill, outline="", width=0)

                # Линии-разделители ячеек: смешиваем заливку с серым/фоном
                # чтобы границы не выглядели резко на цветной заливке
                sep_color = (self._blend(hdr_c, bg, 0.4) if ri == 0
                             else self._blend(fill, "#808080", 0.25))
                canvas.create_line(x0, y1, x1, y1, fill=sep_color, width=1)
                if ci < COLS - 1:
                    canvas.create_line(x1, y0, x1, y1, fill=sep_color, width=1)

                # Имитация текста: одна линия посередине ячейки.
                # В заголовке — толще и цветом hdr_text, в данных — тонкая
                # тёмная линия (или светло-серая на белом фоне)
                mid_y = (y0 + y1) // 2
                tx0, tx1 = x0 + 3, x1 - 3
                line_c = txt_c if ri == 0 else (
                    "#606060" if fill in ("#FFFFFF","#F2F2F2") else
                    self._blend(fill, "#000000", 0.5))
                lw = 2 if ri == 0 else 1
                if tx1 > tx0:
                    canvas.create_line(tx0, mid_y, tx1, mid_y, fill=line_c, width=lw)

        # Внешняя рамка: акцентная если выбран, тонкая серая иначе
        if selected:
            canvas.create_rectangle(PAD - 1, PAD - 1, w - PAD + 1, h - PAD + 1,
                                    outline=t["ACC2"], width=2, fill="")
        else:
            canvas.create_rectangle(PAD - 1, PAD - 1, w - PAD + 1, h - PAD + 1,
                                    outline=t["BORDER"], width=1, fill="")

    @staticmethod
    def _blend(hex1: str, hex2: str, ratio: float) -> str:
        """
        Линейная интерполяция двух hex-цветов.
        ratio=0 → hex1, ratio=1 → hex2, 0.5 → среднее.
        Используется для затемнения/осветления при отрисовке миниатюр.
        """
        try:
            r1,g1,b1 = int(hex1[1:3],16), int(hex1[3:5],16), int(hex1[5:7],16)
            r2,g2,b2 = int(hex2[1:3],16), int(hex2[3:5],16), int(hex2[5:7],16)
            r = int(r1 + (r2-r1)*ratio)
            g = int(g1 + (g2-g1)*ratio)
            b = int(b1 + (b2-b1)*ratio)
            return f"#{r:02X}{g:02X}{b:02X}"
        except Exception:
            return hex1

    def _show_style_picker(self):
        """
        Открывает попап-окно с галереей миниатюр стилей.
        Стили сгруппированы по секциям ("Нет стиля" / Светлые / Средние / Тёмные)
        и располагаются сеткой 4 колонки. Поддерживается прокрутка колесом мыши
        и полосой прокрутки. Выбор сразу сохраняется в settings.json.
        """
        t = self._t()
        names = get_table_style_names(self._lang)

        # overrideredirect = окно без рамки и заголовка (как настоящий dropdown)
        popup = tk.Toplevel(self)
        popup.overrideredirect(True)
        popup.configure(bg=t["CARD"])
        popup.configure(highlightthickness=1, highlightbackground=t["BORDER"])

        # Позиция: прямо под полем выбора стиля, ширина = ширина поля,
        # высота — до 380px или сколько влезет до низа экрана
        self.update_idletasks()
        sf  = self.style_field
        rx  = sf.winfo_rootx()
        ry  = sf.winfo_rooty() + sf.winfo_height()
        pw  = max(sf.winfo_width(), 380)
        screen_h = self.winfo_screenheight()
        popup_h  = min(380, screen_h - ry - 20)
        popup.geometry(f"{pw}x{popup_h}+{rx}+{ry}")

        # ── Прокручиваемый контейнер ──────────────────────────────────────
        # Стандартный tkinter паттерн: Canvas + Scrollbar + внутренний Frame.
        # Frame размещается внутри Canvas через create_window.
        frame_outer = tk.Frame(popup, bg=t["CARD"])
        frame_outer.pack(fill="both", expand=True)

        scrollbar = tk.Scrollbar(frame_outer, orient="vertical",
                                 width=12, bg=t["BG"], troughcolor=t["BG"],
                                 activebackground=t["MUTED"])
        scrollbar.pack(side="right", fill="y")

        scroll_canvas = tk.Canvas(frame_outer, bg=t["CARD"],
                                  highlightthickness=0,
                                  yscrollcommand=scrollbar.set)
        scroll_canvas.pack(side="left", fill="both", expand=True)
        scrollbar.config(command=scroll_canvas.yview)

        inner = tk.Frame(scroll_canvas, bg=t["CARD"])
        inner_id = scroll_canvas.create_window((0, 0), window=inner, anchor="nw")

        # При ресайзе canvas — растягиваем inner на всю ширину
        def _on_canvas_resize(e):
            scroll_canvas.itemconfig(inner_id, width=e.width)
        scroll_canvas.bind("<Configure>", _on_canvas_resize)

        # При изменении содержимого inner — пересчитываем scrollregion
        def _on_inner_resize(e):
            scroll_canvas.configure(scrollregion=scroll_canvas.bbox("all"))
        inner.bind("<Configure>", _on_inner_resize)

        # Колесо мыши прокручивает canvas (bind_all потому что фокус
        # может быть на ребёнке, а не на самом canvas)
        def _on_mousewheel(e):
            scroll_canvas.yview_scroll(int(-1*(e.delta/120)), "units")
        popup.bind_all("<MouseWheel>", _on_mousewheel)

        # ── Параметры миниатюр ───────────────────────────────────────────
        COLS      = 4                                # колонок в сетке
        THUMB_W   = max(60, (pw - 28) // COLS - 8)   # ширина миниатюры
        THUMB_H   = 44                               # высота миниатюры
        LABEL_W   = THUMB_W + 8                      # ширина переноса подписи
        CELL_PAD  = 4

        # ── Названия секций (локализованные) ─────────────────────────────
        groups = [("none", None), ("light", None), ("medium", None), ("dark", None)]
        group_labels = {
            "none":   "",
            "light":  {"ru":"Светлые","en":"Light","de":"Hell",
                       "fr":"Clair","es":"Claro","zh":"浅色",
                       "ar":"فاتح","pt":"Claro"}.get(self._lang,"Light"),
            "medium": {"ru":"Средние","en":"Medium","de":"Mittel",
                       "fr":"Moyen","es":"Medio","zh":"中等",
                       "ar":"متوسط","pt":"Médio"}.get(self._lang,"Medium"),
            "dark":   {"ru":"Тёмные","en":"Dark","de":"Dunkel",
                       "fr":"Foncé","es":"Oscuro","zh":"深色",
                       "ar":"داكن","pt":"Escuro"}.get(self._lang,"Dark"),
        }

        # Словарь key → (canvas, w, h) — нужен для отложенной перерисовки
        # (canvas нельзя нарисовать пока он не получил реальный размер)
        thumb_canvases = {}

        def choose(key, label):
            """Сохраняет выбор и закрывает попап."""
            self._table_style_key = key
            self._settings["table_style_key"] = key
            save_settings(self._settings)
            self.style_var.set(label)
            popup.unbind_all("<MouseWheel>")
            popup.destroy()

        # ── Отрисовка сетки: заголовок секции → ячейки ───────────────────
        current_row = 0
        for group_key, _ in groups:
            # Берём только стили текущей секции, сохраняя их глобальные индексы
            styles_in_group = [(i, s) for i, s in enumerate(ALL_TABLE_STYLES)
                               if s[2] == group_key]
            if not styles_in_group:
                continue

            # Заголовок секции — у "none" не рисуем, она и так одна
            if group_key != "none":
                tk.Label(inner, text=group_labels[group_key],
                         font=(FONT, 8, "bold"),
                         fg=t["MUTED"], bg=t["CARD"],
                         anchor="w", padx=8, pady=4).grid(
                    row=current_row, column=0,
                    columnspan=COLS, sticky="w")
                current_row += 1
                tk.Frame(inner, bg=t["BORDER"], height=1).grid(
                    row=current_row, column=0,
                    columnspan=COLS, sticky="ew", padx=6, pady=(0, 4))
                current_row += 1

            # Раскладываем ячейки секции по сетке слева направо
            col = 0
            for gi, (si, style_tuple) in enumerate(styles_in_group):
                key    = style_tuple[0]
                label  = names[si]
                selected = (key == self._table_style_key)

                cell = tk.Frame(inner, bg=t["CARD"],
                                padx=CELL_PAD, pady=CELL_PAD)
                cell.grid(row=current_row, column=col,
                          padx=2, pady=2, sticky="n")

                cv = tk.Canvas(cell, width=THUMB_W, height=THUMB_H,
                               highlightthickness=0, cursor="hand2",
                               bg=t["CARD"])
                cv.pack()

                lbl = tk.Label(cell, text=label,
                               font=(FONT, 7),
                               fg=t["ACC2"] if selected else t["TEXT"],
                               bg=t["CARD"],
                               wraplength=LABEL_W, justify="center")
                lbl.pack()

                thumb_canvases[key] = (cv, THUMB_W, THUMB_H)

                # ВАЖНО: c=cv, k=key в default-аргументах — стандартный
                # питоновский трюк чтобы зафиксировать значения переменных
                # в момент создания замыкания. Без этого все обработчики
                # ссылались бы на последнее значение из цикла.
                def _enter(e, c=cv, k=key):
                    self._draw_style_thumb(c, k, THUMB_W, THUMB_H, True)
                def _leave(e, c=cv, k=key):
                    self._draw_style_thumb(c, k, THUMB_W, THUMB_H,
                                           k == self._table_style_key)
                def _click(e, k=key, lb=label):
                    choose(k, lb)

                # Биндим обработчики на canvas, label и контейнер ячейки —
                # чтобы клик и hover работали по всей видимой области
                for w in (cv, lbl, cell):
                    w.bind("<Button-1>", _click)
                    w.bind("<Enter>",    _enter)
                    w.bind("<Leave>",    _leave)
                    w.bind("<MouseWheel>", _on_mousewheel)

                col += 1
                if col >= COLS:
                    col = 0
                    current_row += 1

            if col > 0:
                current_row += 1

        # Пустой отступ снизу — иначе последняя строка прижимается к рамке
        tk.Frame(inner, bg=t["CARD"], height=6).grid(
            row=current_row, column=0, columnspan=COLS)

        # Закрываем попап при клике вне его области (перехват через grab_set)
        def close_style_outside(e):
            x, y = e.x_root, e.y_root
            x0, y0 = popup.winfo_rootx(), popup.winfo_rooty()
            x1, y1 = x0 + popup.winfo_width(), y0 + popup.winfo_height()
            # Если координаты клика лежат за пределами окна popup -> закрываем
            if not (x0 <= x <= x1 and y0 <= y <= y1):
                popup.unbind_all("<MouseWheel>")
                popup.destroy()

        popup.bind("<Button-1>", close_style_outside)
        popup.bind("<FocusOut>", lambda e: popup.unbind_all("<MouseWheel>") or popup.destroy() if e.widget == popup else None)

        # ⚠ Важная тонкость: миниатюры нельзя нарисовать прямо сейчас —
        # canvas ещё не получил реальный размер (winfo_width=1).
        # Поэтому рисуем через after() — к моменту вызова окно уже
        # отрисовано и canvas знает свои настоящие размеры.
        def _redraw_all():
            popup.update_idletasks()
            for k, (cv, tw, th) in thumb_canvases.items():
                self._draw_style_thumb(cv, k, tw, th, k == self._table_style_key)
            scroll_canvas.update_idletasks()

        popup.after(15, _redraw_all)
        popup.focus_set()
        popup.grab_set()


    # ═════════════════════════════════════════════════════════════════════════
    # Автозагрузка и подсказки
    # ═════════════════════════════════════════════════════════════════════════

    def _auto_load(self):
        """
        При запуске пытаемся подхватить файлы рядом с .exe:
          • первый .csv → в поле "CSV"
          • columns.txt → в поле "Фильтр колонок"
        Это удобно когда .exe и файлы лежат в одной папке —
        пользователю не нужно ничего выбирать вручную.
        """
        # Текст в drop-zone выставляется здесь — раньше нельзя было,
        # т.к. self.s() требует уже инициализированный _lang
        if not self._dz_text:
            self._dz_text = self.s("dz_idle")
            self._redraw_dz()
        if not self.csv_var.get():
            p = find_csv_in_exe_dir()
            if p:
                self._load_csv(p)
        if not self.col_var.get():
            auto = find_columns_txt()
            if auto:
                self.col_var.set(auto)
                self._refresh_hint()

    # ── Hint ──────────────────────────────────────────────────────────────────

    def _refresh_hint(self):
        t = self._t()
        p = self.col_var.get().strip()
        if p and Path(p).exists():
            parsed = parse_columns_txt(p)
            n_cols   = len(parsed["columns"])
            n_strips = len(parsed["strip_prefixes"])
            parts = []
            if n_cols:
                parts.append(f"{n_cols}{self.s('hint_cols')}") 
            else:
                parts.append(self.s("hint_all"))
            if n_strips:
                parts.append(f"{n_strips}{self.s('hint_strips')}") 
            self.hint_var.set(self.s("hint_found") + ", ".join(parts))
            self.hint_lbl.config(fg=GREEN)
        else:
            self.hint_var.set(self.s("hint_none"))
            self.hint_lbl.config(fg=t["MUTED"])

    # ═════════════════════════════════════════════════════════════════════════
    # Статус и прогресс-бар
    # ═════════════════════════════════════════════════════════════════════════

    def _set_status(self, text, color=None):
        """Меняет текст статуса. Если color не задан — цвет MUTED из темы."""
        t = self._t()
        self.status_var.set(text)
        self.status_lbl.config(fg=color or t["MUTED"])

    def _start_pb(self):
        """Запускает бесконечную анимацию прогресс-бара (бегущий сегмент)."""
        self._pb_anim = True
        self._pb_pos  = 0
        self._animate_pb()

    def _animate_pb(self):
        """
        Кадр анимации прогресс-бара. Рисует "бегущий" сегмент длиной ~25%
        от ширины бара. Перевызывается через after() пока _pb_anim=True.
        """
        if not self._pb_anim:
            return
        w   = self._pb_w
        seg = max(w // 4, 60)               # минимум 60px чтобы был виден
        x0  = self._pb_pos % w
        x1  = min(x0 + seg, w)
        self.pb_canvas.coords(self.pb_rect, x0, 0, x1, 6)
        self._pb_pos = (self._pb_pos + 6) % w
        self.after(25, self._animate_pb)    # ~40 кадров/с

    def _stop_pb(self, success=True):
        """
        Останавливает анимацию и заполняет бар целиком.
        Зелёный цвет = успех, красный = ошибка.
        """
        self._pb_anim = False
        self.pb_canvas.coords(self.pb_rect, 0, 0, self._pb_w, 6)
        self.pb_canvas.itemconfig(self.pb_rect,
                                  fill=GREEN if success else RED)

# ═════════════════════════════════════════════════════════════════════════════
    # Запуск конвертации
    # ═════════════════════════════════════════════════════════════════════════════

    def _run(self):
        """
        Обработчик кнопки "Преобразовать". Валидирует входные данные,
        запускает конвертацию в отдельном потоке (чтобы UI не подвисал)
        и через self.after() планирует обновление UI после завершения.
        """
        xlsx_path = self.xlsx_var.get().strip()

        # Валидация списка файлов
        if not hasattr(self, '_csv_paths') or not self._csv_paths:
            messagebox.showwarning(self.s("err_ttl"), self.s("no_file"))
            return
            
        for p in self._csv_paths:
            if not os.path.isfile(p):
                messagebox.showerror(self.s("err_ttl"), f"{self.s('file_not_found')}\n{p}")
                return
                
        # Проверяем путь сохранения
        if not xlsx_path:
            messagebox.showwarning(self.s("err_ttl"), self.s("no_path"))
            return

        col_path   = self.col_var.get().strip()
        style_name = TABLE_STYLE_BY_KEY.get(self._table_style_key)
        row_h      = self.rh_var.get().strip() # Забираем высоту строк
        dlm        = self.dlm_var.get().strip() # Забираем разделитель
        
        is_to_csv = getattr(self, '_mode', 'to_xlsx') == 'to_csv'

        # Вложенная функция, которая запускается ТОЛЬКО после подтверждения разделителя
        def start_conversion(final_dlm):
            # UI в режим "работаем"
            self.btn.config(state="disabled")
            self._set_status(self.s("status_working"), self._t()["ACC2"])
            self._start_pb()
            
            # Конвертация в фоне — даже на больших файлах UI остаётся отзывчивым.
            # Результат приходит через self.after(0, ...) — это безопасный
            # способ выполнить код в главном (UI) потоке из фонового.
            def worker():
                try:
                    if is_to_csv:
                        # Обратная конвертация XLSX -> CSV
                        rows, cols = convert_xlsx_to_csv(self._csv_paths, xlsx_path, delimiter=final_dlm)
                    else:
                        # Стандартная конвертация CSV -> XLSX
                        rows, cols = convert(self._csv_paths, xlsx_path, col_path, style_name, row_height=row_h, delimiter=final_dlm)
                        
                    self.after(0, lambda: self._done(True, rows, cols, xlsx_path))
                except Exception as e:
                    self.after(0, lambda: self._done(False, error=str(e)))

            threading.Thread(target=worker, daemon=True).start()

        # Если делаем CSV для Jira -> сначала показываем диалог выбора разделителя!
        if is_to_csv:
            self._show_delimiter_dialog(start_conversion)
        else:
            start_conversion(dlm)
            
    def _show_delimiter_dialog(self, on_confirm):
        """Модальное окно выбора разделителя перед сохранением в CSV."""
        t = self._t()
        win = tk.Toplevel(self)
        
        win.title(self.s("dlm_ttl"))
        win.configure(bg=t["CARD"])
        win.resizable(False, False)
        win.transient(self) # связан с главным окном
        win.grab_set()      # модальный — блокирует главное окно

        tk.Label(win, text=self.s("dlm_msg"), font=(FONT, 10, "bold"), fg=t["TEXT"], bg=t["CARD"]).pack(pady=(20, 10), padx=30)

        var = tk.StringVar(value=self.dlm_var.get())
        frame = tk.Frame(win, bg=t["CARD"])
        frame.pack(pady=(0, 20), anchor="w", padx=40)

        # Берем названия разделителей прямо из словаря локализации
        options = [
            (";", self.s("dlm_semi")), 
            (",", self.s("dlm_comma")), 
            ("|", self.s("dlm_pipe")), 
            ("Tab", self.s("dlm_tab"))
        ]
        
        for d, name in options:
            tk.Radiobutton(frame, text=name, variable=var, value=d,
                           font=(FONT, 9), fg=t["TEXT"], bg=t["CARD"],
                           selectcolor=t["CARD"], activebackground=t["CARD"], cursor="hand2").pack(anchor="w", pady=2)

        btn_frame = tk.Frame(win, bg=t["CARD"])
        btn_frame.pack(fill="x", padx=20, pady=(0, 20))

        def on_ok():
            dlm = var.get()
            self._settings["csv_delimiter"] = dlm
            save_settings(self._settings)
            self.dlm_var.set(dlm) # Синхронизируем с главным окном
            win.destroy()
            on_confirm(dlm) # Запускаем конвертацию

        # Кнопка Продолжить (зеленая)
        tk.Button(btn_frame, text=self.s("dlm_ok"), font=(FONT, 9), fg="white", bg=t["ACC2"],
                  activebackground="#1A4A8A", activeforeground="white", relief="flat", bd=0, padx=15, pady=6,
                  cursor="hand2", command=on_ok).pack(side="left", padx=(10, 0))
                  
        # Кнопка Закрыть
        tk.Button(btn_frame, text=self.s("close"), font=(FONT, 9), fg=t["TEXT"], bg=t["ENTRY"],
                  activebackground=t["BORDER"], activeforeground=t["TEXT"], relief="flat", bd=0, padx=15, pady=6,
                  cursor="hand2", command=win.destroy).pack(side="right", padx=(0, 10))

        # Центрируем и красим заголовок в темную/светлую тему
        win.update_idletasks()
        pw, ph = self.winfo_x(), self.winfo_y()
        ww = self.winfo_width()
        win.geometry(f"+{pw + ww//2 - win.winfo_reqwidth()//2}+{ph + 60}")
        self.after(50, lambda: self._apply_title_theme(win))

    def _show_done_dialog(self, path):
        """Кастомное диалоговое окно об успешном завершении с выбором действий."""
        t = self._t()
        win = tk.Toplevel(self)
        win.title(self.s("done_ttl"))
        win.configure(bg=t["CARD"])
        win.resizable(False, False)
        win.transient(self)
        win.grab_set()

        # Очищаем сообщение от старого вопроса "Открыть папку?"
        raw_msg = self.s("done_msg").format(path=path)
        clean_msg = raw_msg.split("\n\n")[0] # Оставит только "Файл сохранен: Путь"

        tk.Label(win, text=clean_msg, font=(FONT, 10), fg=t["TEXT"], bg=t["CARD"],
                 justify="left").pack(pady=(20, 15), padx=20)

        btn_frame = tk.Frame(win, bg=t["CARD"])
        btn_frame.pack(fill="x", padx=20, pady=(0, 20))

        # Локализация кнопок "на лету" (чтобы не переписывать весь огромный словарь)
        loc_file = {"ru": "Открыть файл", "en": "Open file", "de": "Datei öffnen",
                    "fr": "Ouvrir le fichier", "es": "Abrir archivo", "zh": "打开文件",
                    "ar": "فتح الملف", "pt": "Abrir ficheiro"}.get(self._lang, "Open file")
        loc_folder = {"ru": "Открыть папку", "en": "Open folder", "de": "Ordner öffnen",
                      "fr": "Ouvrir le dossier", "es": "Abrir carpeta", "zh": "打开文件夹",
                      "ar": "فتح المجلد", "pt": "Abrir pasta"}.get(self._lang, "Open folder")

        def on_file():
            os.startfile(path)
            win.destroy()

        def on_folder():
            os.startfile(str(Path(path).parent))
            win.destroy()

        # Кнопка: Открыть файл (Зеленая)
        tk.Button(btn_frame, text=loc_file, font=(FONT, 9), fg="white", bg="#4A7C59",
                  activebackground="#2E5E3E", activeforeground="white", relief="flat",
                  bd=0, padx=12, pady=6, cursor="hand2", command=on_file).pack(side="left", padx=(0, 10))

        # Кнопка: Открыть папку (Синяя)
        tk.Button(btn_frame, text=loc_folder, font=(FONT, 9), fg="white", bg=t["ACC2"],
                  activebackground="#1A4A8A", activeforeground="white", relief="flat",
                  bd=0, padx=12, pady=6, cursor="hand2", command=on_folder).pack(side="left", padx=(0, 10))

        # Кнопка: Закрыть (Серая/Нейтральная)
        tk.Button(btn_frame, text=self.s("close"), font=(FONT, 9), fg=t["TEXT"], bg=t["ENTRY"],
                  activebackground=t["BORDER"], activeforeground=t["TEXT"], relief="flat",
                  bd=0, padx=12, pady=6, cursor="hand2", command=win.destroy).pack(side="right")

        win.update_idletasks()
        pw, ph = self.winfo_x(), self.winfo_y()
        ww = self.winfo_width()
        win.geometry(f"+{pw + ww//2 - win.winfo_reqwidth()//2}+{ph + 60}")
        self.after(50, lambda: self._apply_title_theme(win))

    def _done(self, ok, rows=0, cols=0, path="", error=""):
        """Колбэк завершения конвертации. Возвращает UI в исходное состояние."""
        self._stop_pb(ok)
        self.btn.config(state="normal")
        if ok:
            self._set_status(
                self.s("done_status").format(rows=rows, cols=cols, name=Path(path).name),
                GREEN)
            # Вызываем наше кастомное окно вместо стандартного
            self._show_done_dialog(path)
        else:
            self._set_status(self.s("err_status").format(error=error), RED)
            messagebox.showerror(self.s("err_ttl"), error)


# ═════════════════════════════════════════════════════════════════════════════
# Точка входа
# ═════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    # При запуске .exe (PyInstaller onefile) рабочая директория может
    # указывать на временную папку распаковки, а не на папку с .exe.
    # Принудительно меняем её — это нужно для load_settings(), find_csv_in_exe_dir()
    # и других функций которые ожидают правильный CWD.
    if getattr(sys, "frozen", False):
        os.chdir(Path(sys.executable).resolve().parent)

    import ctypes

    # Создаем уникальный идентификатор (Mutex) для нашей программы
    mutex_name = "JiraCSV_XLSX_Converter_Mutex_Maxsteff"
    mutex = ctypes.windll.kernel32.CreateMutexW(None, False, mutex_name)

    # Ошибка 183 (ERROR_ALREADY_EXISTS) означает, что программа уже запущена
    if ctypes.windll.kernel32.GetLastError() == 183:
        hwnd = 0
        # Проверяем заголовки программы на всех языках, чтобы найти окно
        for lang_dict in STRINGS.values():
            title = lang_dict.get("title", APP_NAME)
            hwnd = ctypes.windll.user32.FindWindowW(None, title)
            if hwnd:
                break
        
        if hwnd:
            # SW_RESTORE = 9 (восстанавливает окно, если оно свернуто в панель задач)
            ctypes.windll.user32.ShowWindow(hwnd, 9)
            # Выводит найденное окно на передний план
            ctypes.windll.user32.SetForegroundWindow(hwnd)
            
        sys.exit(0) # Моментально закрываем копию программы

    app = App()
    app.mainloop()