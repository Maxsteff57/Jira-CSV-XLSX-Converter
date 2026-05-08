"""
Тесты ядра конвертера Jira CSV → XLSX.

Покрываем:
  • parse_jira_date       — парсинг дат из экспорта Jira
  • auto_rename           — очистка имён колонок
  • parse_columns_txt     — разбор файла фильтра с секциями [columns]/[strip]
  • load_columns_filter   — загрузка фильтра (с явным путём и без)
  • convert               — полная конвертация CSV → XLSX
  • convert_xlsx_to_csv   — обратная конвертация XLSX → CSV
  • generate_columns_example — генерация примера columns.txt
  • ALL_TABLE_STYLES       — корректность палитры стилей Excel
  • STRINGS               — наличие обязательных ключей во всех языках
"""
import os
import sys
import json
import pytest
import tempfile
import textwrap
from pathlib import Path

# Добавляем корень проекта в sys.path чтобы импортировать app
sys.path.insert(0, str(Path(__file__).parent.parent))

# ── Мокаем winreg и tkinter до импорта app ───────────────────────────────────
# На Linux / CI этих модулей нет, а GUI нам запускать не нужно.
from unittest.mock import MagicMock, patch
sys.modules.setdefault("winreg", MagicMock())
sys.modules.setdefault("tkinter", MagicMock())
sys.modules.setdefault("tkinter.filedialog", MagicMock())
sys.modules.setdefault("tkinter.messagebox", MagicMock())
sys.modules.setdefault("tkinter.ttk", MagicMock())
sys.modules.setdefault("tkinterdnd2", MagicMock())

import pandas as pd
from openpyxl import load_workbook

import app as core


# ═════════════════════════════════════════════════════════════════════════════
# Fixtures
# ═════════════════════════════════════════════════════════════════════════════

@pytest.fixture
def tmp(tmp_path):
    """Возвращает временную директорию как Path."""
    return tmp_path


@pytest.fixture
def simple_csv(tmp_path) -> Path:
    """Минимальный CSV в формате Jira: три колонки, три строки данных."""
    p = tmp_path / "test.csv"
    p.write_text(
        "Ключ проблемы;Тема;Пользовательское поле (Статус)\n"
        "TASK-1;Первая задача;Открыта\n"
        "TASK-2;Вторая задача;В работе\n"
        "TASK-3;Третья задача;Закрыта\n",
        encoding="utf-8-sig",
    )
    return p


@pytest.fixture
def date_csv(tmp_path) -> Path:
    """CSV с датами в формате Jira."""
    p = tmp_path / "dates.csv"
    p.write_text(
        "Ключ;Дата регистрации замечания;Плановая дата устранения\n"
        "T-1;05/фев/26 12:00 AM;13/мар/26\n"
        "T-2;01/янв/2026;31/дек/25 11:59 PM\n"
        "T-3;нет даты;;\n",
        encoding="utf-8-sig",
    )
    return p


@pytest.fixture
def columns_txt(tmp_path) -> Path:
    """Файл columns.txt с обеими секциями."""
    p = tmp_path / "columns.txt"
    p.write_text(
        textwrap.dedent("""\
            # комментарий игнорируется

            [columns]
            Ключ проблемы
            Тема
            Статус

            [strip]
            Пользовательское поле\x20
            Custom field\x20
        """),
        encoding="utf-8",
    )
    return p


# ═════════════════════════════════════════════════════════════════════════════
# parse_jira_date
# ═════════════════════════════════════════════════════════════════════════════

class TestParseJiraDate:

    def test_russian_month_short_year(self):
        assert core.parse_jira_date("05/фев/26 12:00 AM") == "05.02.2026"

    def test_russian_month_full_year(self):
        assert core.parse_jira_date("01/янв/2026") == "01.01.2026"

    def test_russian_month_no_time(self):
        assert core.parse_jira_date("31/дек/25") == "31.12.2025"

    def test_russian_may_alternative_form(self):
        # "мая" — родительный падеж, встречается в некоторых локалях Jira
        assert core.parse_jira_date("07/мая/26") == "07.05.2026"

    def test_english_month(self):
        assert core.parse_jira_date("15/Jan/26") == "15.01.2026"

    def test_day_padding(self):
        assert core.parse_jira_date("3/мар/26") == "03.03.2026"

    def test_empty_string_passthrough(self):
        assert core.parse_jira_date("") == ""

    def test_non_date_string_passthrough(self):
        assert core.parse_jira_date("не дата") == "не дата"

    def test_none_passthrough(self):
        assert core.parse_jira_date(None) is None

    def test_number_passthrough(self):
        assert core.parse_jira_date(42) == 42

    def test_time_is_stripped(self):
        # Время должно отбрасываться
        result = core.parse_jira_date("12/авг/26 11:59 PM")
        assert result == "12.08.2026"

    def test_all_12_russian_months(self):
        months = [
            ("янв", "01"), ("фев", "02"), ("мар", "03"), ("апр", "04"),
            ("май", "05"), ("июн", "06"), ("июл", "07"), ("авг", "08"),
            ("сен", "09"), ("окт", "10"), ("ноя", "11"), ("дек", "12"),
        ]
        for abbr, num in months:
            result = core.parse_jira_date(f"01/{abbr}/26")
            assert result == f"01.{num}.2026", f"Failed for month {abbr!r}"


# ═════════════════════════════════════════════════════════════════════════════
# auto_rename
# ═════════════════════════════════════════════════════════════════════════════

class TestAutoRename:

    def test_builtin_jira_pattern(self):
        assert core.auto_rename("Пользовательское поле (Статус)") == "Статус"

    def test_builtin_pattern_long_name(self):
        r = core.auto_rename("Пользовательское поле (Дата регистрации замечания)")
        assert r == "Дата регистрации замечания"

    def test_plain_column_unchanged(self):
        assert core.auto_rename("Тема") == "Тема"
        assert core.auto_rename("Ключ проблемы") == "Ключ проблемы"

    def test_custom_prefix_with_parens(self):
        r = core.auto_rename("Custom field (Priority)", ["Custom field "])
        assert r == "Priority"

    def test_custom_prefix_without_parens(self):
        r = core.auto_rename("Custom field Priority", ["Custom field "])
        assert r == "Priority"

    def test_custom_prefix_overrides_builtin(self):
        # Если задан [strip] с кастомным префиксом, он проверяется первым
        r = core.auto_rename(
            "Пользовательское поле (Тема)",
            ["Пользовательское поле "]
        )
        assert r == "Тема"

    def test_unrecognized_with_custom_prefix(self):
        # Колонка не начинается ни с одного префикса → возвращается как есть
        r = core.auto_rename("Обычная колонка", ["Custom field "])
        assert r == "Обычная колонка"

    def test_whitespace_stripped(self):
        # Пробелы вокруг имени колонки должны убираться
        assert core.auto_rename("  Тема  ") == "Тема"

    def test_empty_strip_list_uses_builtin(self):
        r = core.auto_rename("Пользовательское поле (Статус)", [])
        assert r == "Статус"


# ═════════════════════════════════════════════════════════════════════════════
# parse_columns_txt
# ═════════════════════════════════════════════════════════════════════════════

class TestParseColumnsTxt:

    def test_basic_columns_only(self, tmp):
        p = tmp / "c.txt"
        p.write_text("[columns]\nТема\nСтатус\nКлюч проблемы\n", encoding="utf-8")
        result = core.parse_columns_txt(str(p))
        assert result["columns"] == ["Тема", "Статус", "Ключ проблемы"]
        assert result["strip_prefixes"] == []

    def test_strip_section(self, tmp):
        p = tmp / "c.txt"
        p.write_text(
            "[columns]\nТема\n\n[strip]\nПользовательское поле \nCustom field \n",
            encoding="utf-8",
        )
        result = core.parse_columns_txt(str(p))
        assert result["columns"] == ["Тема"]
        assert result["strip_prefixes"] == ["Пользовательское поле ", "Custom field "]

    def test_trailing_space_preserved_in_strip(self, tmp):
        """Пробел в конце строки [strip] — это часть префикса."""
        p = tmp / "c.txt"
        p.write_text("[strip]\nПользовательское поле \n", encoding="utf-8")
        result = core.parse_columns_txt(str(p))
        assert result["strip_prefixes"][0].endswith(" ")

    def test_comments_ignored(self, tmp):
        p = tmp / "c.txt"
        p.write_text("# заголовок\n[columns]\n# пропустить\nТема\n", encoding="utf-8")
        result = core.parse_columns_txt(str(p))
        assert result["columns"] == ["Тема"]

    def test_empty_lines_ignored(self, tmp):
        p = tmp / "c.txt"
        p.write_text("\n\n[columns]\n\nТема\n\nСтатус\n\n", encoding="utf-8")
        result = core.parse_columns_txt(str(p))
        assert result["columns"] == ["Тема", "Статус"]

    def test_nonexistent_file_returns_empty(self):
        result = core.parse_columns_txt("/nonexistent/path/columns.txt")
        assert result == {"columns": [], "strip_prefixes": []}

    def test_empty_path_returns_empty(self):
        result = core.parse_columns_txt("")
        assert result == {"columns": [], "strip_prefixes": []}

    def test_default_section_is_columns(self, tmp):
        """Без явного [columns] строки всё равно попадают в секцию columns."""
        p = tmp / "c.txt"
        p.write_text("Тема\nСтатус\n", encoding="utf-8")
        result = core.parse_columns_txt(str(p))
        assert result["columns"] == ["Тема", "Статус"]

    def test_utf8_bom_handled(self, tmp):
        p = tmp / "c.txt"
        # Записываем с BOM (utf-8-sig) как это делает Notepad на Windows
        p.write_bytes(b"\xef\xbb\xbf[columns]\n\xd0\xa2\xd0\xb5\xd0\xbc\xd0\xb0\n")
        result = core.parse_columns_txt(str(p))
        assert result["columns"] == ["Тема"]

    def test_full_file(self, columns_txt):
        result = core.parse_columns_txt(str(columns_txt))
        assert result["columns"] == ["Ключ проблемы", "Тема", "Статус"]
        assert "Пользовательское поле " in result["strip_prefixes"]
        assert "Custom field " in result["strip_prefixes"]


# ═════════════════════════════════════════════════════════════════════════════
# load_columns_filter
# ═════════════════════════════════════════════════════════════════════════════

class TestLoadColumnsFilter:

    def test_explicit_path(self, columns_txt):
        keep, strips = core.load_columns_filter(str(columns_txt))
        assert keep == ["Ключ проблемы", "Тема", "Статус"]
        assert "Пользовательское поле " in strips

    def test_no_file_returns_none(self, tmp):
        keep, strips = core.load_columns_filter(str(tmp / "nonexistent.txt"))
        assert keep is None
        assert strips == []

    def test_empty_columns_section_returns_none(self, tmp):
        p = tmp / "c.txt"
        p.write_text("[columns]\n# только комментарии\n", encoding="utf-8")
        keep, strips = core.load_columns_filter(str(p))
        assert keep is None  # пустой список → None


# ═════════════════════════════════════════════════════════════════════════════
# convert  (CSV → XLSX)
# ═════════════════════════════════════════════════════════════════════════════

class TestConvert:

    def test_basic_conversion(self, simple_csv, tmp):
        out = str(tmp / "out.xlsx")
        rows, cols = core.convert([str(simple_csv)], out)
        assert rows == 3
        assert cols == 3
        assert Path(out).exists()

    def test_output_has_correct_headers(self, simple_csv, tmp):
        out = str(tmp / "out.xlsx")
        core.convert([str(simple_csv)], out)
        wb = load_workbook(out)
        ws = wb.active
        headers = [ws.cell(1, c).value for c in range(1, 4)]
        # Заголовок "Пользовательское поле (Статус)" должен стать "Статус"
        assert "Статус" in headers
        assert "Тема" in headers
        assert "Ключ проблемы" in headers

    def test_jira_prefix_stripped_in_headers(self, simple_csv, tmp):
        out = str(tmp / "out.xlsx")
        core.convert([str(simple_csv)], out)
        wb = load_workbook(out)
        ws = wb.active
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        assert not any("Пользовательское поле" in str(h) for h in headers)

    def test_column_filter(self, simple_csv, tmp, columns_txt):
        out = str(tmp / "out.xlsx")
        rows, cols = core.convert([str(simple_csv)], out, columns_path=str(columns_txt))
        assert cols == 3  # только Ключ проблемы, Тема, Статус

    def test_data_row_count(self, simple_csv, tmp):
        out = str(tmp / "out.xlsx")
        rows, _ = core.convert([str(simple_csv)], out)
        wb = load_workbook(out)
        ws = wb.active
        assert ws.max_row == rows + 1  # +1 заголовок

    def test_date_columns_converted(self, date_csv, tmp):
        out = str(tmp / "out.xlsx")
        core.convert([str(date_csv)], out)
        wb = load_workbook(out)
        ws = wb.active
        # Находим колонку с "дата" в имени
        date_col = None
        for c in range(1, ws.max_column + 1):
            if "дата" in str(ws.cell(1, c).value or "").lower():
                date_col = c
                break
        assert date_col is not None
        # Первое значение должно быть в формате ДД.ММ.ГГГГ
        val = ws.cell(2, date_col).value
        assert val == "05.02.2026", f"Got: {val!r}"

    def test_table_style_applied(self, simple_csv, tmp):
        out = str(tmp / "out.xlsx")
        core.convert([str(simple_csv)], out, table_style="TableStyleMedium2")
        wb = load_workbook(out)
        ws = wb.active
        # В xlsx должна быть хотя бы одна таблица
        assert len(ws.tables) > 0

    def test_no_style_no_table(self, simple_csv, tmp):
        out = str(tmp / "out.xlsx")
        core.convert([str(simple_csv)], out, table_style=None)
        wb = load_workbook(out)
        ws = wb.active
        assert len(ws.tables) == 0

    def test_freeze_panes_set(self, simple_csv, tmp):
        out = str(tmp / "out.xlsx")
        core.convert([str(simple_csv)], out)
        wb = load_workbook(out)
        ws = wb.active
        assert ws.freeze_panes == "A2"

    def test_multiple_csv_merged(self, tmp):
        """Два CSV с одинаковой схемой должны склеиваться."""
        csv1 = tmp / "a.csv"
        csv2 = tmp / "b.csv"
        csv1.write_text("Ключ;Тема\nT-1;Первая\nT-2;Вторая\n", encoding="utf-8-sig")
        csv2.write_text("Ключ;Тема\nT-3;Третья\nT-4;Четвёртая\n", encoding="utf-8-sig")
        out = str(tmp / "merged.xlsx")
        rows, cols = core.convert([str(csv1), str(csv2)], out)
        assert rows == 4
        assert cols == 2

    def test_duplicates_removed_on_merge(self, tmp):
        """При склейке дублирующиеся строки убираются."""
        csv1 = tmp / "a.csv"
        csv2 = tmp / "b.csv"
        csv1.write_text("Ключ;Тема\nT-1;Задача\n", encoding="utf-8-sig")
        csv2.write_text("Ключ;Тема\nT-1;Задача\n", encoding="utf-8-sig")
        out = str(tmp / "dedup.xlsx")
        rows, _ = core.convert([str(csv1), str(csv2)], out)
        assert rows == 1

    def test_empty_list_returns_zero(self, tmp):
        out = str(tmp / "empty.xlsx")
        rows, cols = core.convert([], out)
        assert rows == 0
        assert cols == 0

    def test_row_height_applied(self, simple_csv, tmp):
        out = str(tmp / "out.xlsx")
        core.convert([str(simple_csv)], out, row_height="30")
        wb = load_workbook(out)
        ws = wb.active
        # Хотя бы одна строка данных должна иметь заданную высоту
        assert ws.row_dimensions[2].height == 30.0

    def test_comma_delimiter(self, tmp):
        p = tmp / "comma.csv"
        p.write_text("Key,Subject\nT-1,Task\n", encoding="utf-8-sig")
        out = str(tmp / "out.xlsx")
        rows, cols = core.convert([str(p)], out, delimiter=",")
        assert rows == 1
        assert cols == 2


# ═════════════════════════════════════════════════════════════════════════════
# convert_xlsx_to_csv  (XLSX → CSV)
# ═════════════════════════════════════════════════════════════════════════════

class TestConvertXlsxToCsv:

    def _make_xlsx(self, path: Path, data: list[list]) -> Path:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        for row in data:
            ws.append(row)
        wb.save(str(path))
        return path

    def test_basic_round_trip(self, tmp):
        xls = self._make_xlsx(tmp / "data.xlsx",
                              [["Ключ", "Тема"], ["T-1", "Задача"]])
        out = str(tmp / "out.csv")
        rows, cols = core.convert_xlsx_to_csv([str(xls)], out)
        assert rows == 1
        assert cols == 2
        content = Path(out).read_text(encoding="utf-8")
        assert "Ключ" in content
        assert "T-1" in content

    def test_default_semicolon_delimiter(self, tmp):
        xls = self._make_xlsx(tmp / "d.xlsx", [["A", "B"], ["1", "2"]])
        out = str(tmp / "out.csv")
        core.convert_xlsx_to_csv([str(xls)], out)
        content = Path(out).read_text(encoding="utf-8")
        assert ";" in content

    def test_comma_delimiter(self, tmp):
        xls = self._make_xlsx(tmp / "d.xlsx", [["A", "B"], ["1", "2"]])
        out = str(tmp / "out.csv")
        core.convert_xlsx_to_csv([str(xls)], out, delimiter=",")
        content = Path(out).read_text(encoding="utf-8")
        assert "A,B" in content

    def test_multiple_xlsx_merged(self, tmp):
        x1 = self._make_xlsx(tmp / "a.xlsx", [["K"], ["T-1"], ["T-2"]])
        x2 = self._make_xlsx(tmp / "b.xlsx", [["K"], ["T-3"]])
        out = str(tmp / "out.csv")
        rows, _ = core.convert_xlsx_to_csv([str(x1), str(x2)], out)
        assert rows == 3

    def test_empty_list_returns_zero(self, tmp):
        out = str(tmp / "empty.csv")
        rows, cols = core.convert_xlsx_to_csv([], out)
        assert rows == 0
        assert cols == 0


# ═════════════════════════════════════════════════════════════════════════════
# generate_columns_example
# ═════════════════════════════════════════════════════════════════════════════

class TestGenerateColumnsExample:

    def test_creates_file(self, simple_csv, tmp):
        result = core.generate_columns_example(str(simple_csv), tmp)
        assert result != ""
        assert Path(result).exists()

    def test_file_named_columns_txt(self, simple_csv, tmp):
        result = core.generate_columns_example(str(simple_csv), tmp)
        assert Path(result).name == "columns.txt"

    def test_contains_column_names(self, simple_csv, tmp):
        result = core.generate_columns_example(str(simple_csv), tmp)
        content = Path(result).read_text(encoding="utf-8")
        assert "Ключ проблемы" in content
        assert "Тема" in content
        # "Пользовательское поле (Статус)" должен быть переименован в "Статус"
        assert "Статус" in content
        # В строках данных (не комментариях) не должно быть "Пользовательское поле"
        data_lines = [l for l in content.splitlines()
                      if l.strip() and not l.strip().startswith("#")]
        assert not any("Пользовательское поле" in l for l in data_lines)

    def test_has_columns_section(self, simple_csv, tmp):
        result = core.generate_columns_example(str(simple_csv), tmp)
        content = Path(result).read_text(encoding="utf-8")
        assert "[columns]" in content

    def test_has_strip_hint(self, simple_csv, tmp):
        result = core.generate_columns_example(str(simple_csv), tmp)
        content = Path(result).read_text(encoding="utf-8")
        assert "[strip]" in content

    def test_custom_strip_prefixes_applied(self, tmp):
        p = tmp / "in.csv"
        p.write_text("Custom field (Priority);Custom field (Status)\n", encoding="utf-8-sig")
        result = core.generate_columns_example(str(p), tmp, strip_prefixes=["Custom field "])
        content = Path(result).read_text(encoding="utf-8")
        assert "Priority" in content
        assert "Status" in content
        assert "Custom field" not in content.replace("Custom field ", "")

    def test_invalid_csv_returns_empty(self, tmp):
        result = core.generate_columns_example("/nonexistent/file.csv", tmp)
        assert result == ""


# ═════════════════════════════════════════════════════════════════════════════
# ALL_TABLE_STYLES — структура и корректность палитры
# ═════════════════════════════════════════════════════════════════════════════

class TestAllTableStyles:

    STYLE_KEYS = [s[0] for s in core.ALL_TABLE_STYLES]
    XLSX_NAMES = [s[1] for s in core.ALL_TABLE_STYLES if s[1]]

    def test_no_duplicate_keys(self):
        assert len(self.STYLE_KEYS) == len(set(self.STYLE_KEYS))

    def test_no_duplicate_xlsx_names(self):
        assert len(self.XLSX_NAMES) == len(set(self.XLSX_NAMES))

    def test_all_groups_valid(self):
        valid = {"none", "light", "medium", "dark"}
        for s in core.ALL_TABLE_STYLES:
            assert s[2] in valid, f"Неверная группа у {s[0]!r}: {s[2]!r}"

    def test_colors_are_hex(self):
        import re
        pat = re.compile(r"^#[0-9A-Fa-f]{6}$")
        for s in core.ALL_TABLE_STYLES:
            for color in s[3:]:  # hdr, odd, even, hdr_text
                assert pat.match(color), f"Не HEX у {s[0]!r}: {color!r}"

    def test_by_key_dict_complete(self):
        assert set(core.TABLE_STYLE_BY_KEY) == set(self.STYLE_KEYS)

    def test_colors_dict_complete(self):
        assert set(core.TABLE_STYLE_COLORS) == set(self.STYLE_KEYS)

    def test_known_palette_blue_light(self):
        # Office 2007 синий = #4F81BD
        colors = core.TABLE_STYLE_COLORS["s_l_blue"]
        assert colors[0] == "#4F81BD"

    def test_known_palette_green_medium(self):
        colors = core.TABLE_STYLE_COLORS["s_m_green"]
        assert colors[0] == "#9BBB59"

    def test_known_palette_dark_orange_structure(self):
        # У тёмных стилей: header < odd < even по насыщенности (от тёмного к светлому)
        colors = core.TABLE_STYLE_COLORS["s_d_orange"]
        # header должен быть темнее odd
        hdr_r = int(colors[0][1:3], 16)
        odd_r = int(colors[1][1:3], 16)
        assert hdr_r < odd_r  # тёмный заголовок → меньший красный канал

    def test_none_style_has_no_xlsx_name(self):
        assert core.TABLE_STYLE_BY_KEY["s_none"] is None


# ═════════════════════════════════════════════════════════════════════════════
# STRINGS — локализация
# ═════════════════════════════════════════════════════════════════════════════

class TestStrings:

    REQUIRED_KEYS = [
        "title", "csv_lbl", "xlsx_lbl", "col_lbl", "style_lbl",
        "convert_btn", "browse", "status_idle", "done_status",
        "done_msg", "err_status", "about_desc", "about_tg",
        "dz_idle", "hint_none",
        "s_none", "s_l_blue", "s_m_green", "s_d_blue",
    ]
    LANGUAGES = ["ru", "en", "de", "fr", "es", "zh", "ar", "pt"]

    def test_all_languages_present(self):
        for lang in self.LANGUAGES:
            assert lang in core.STRINGS, f"Язык {lang!r} отсутствует в STRINGS"

    @pytest.mark.parametrize("lang", LANGUAGES)
    def test_required_keys_present(self, lang):
        missing = [k for k in self.REQUIRED_KEYS if k not in core.STRINGS[lang]]
        assert not missing, f"Язык {lang!r}: отсутствуют ключи {missing}"

    @pytest.mark.parametrize("lang", LANGUAGES)
    def test_no_self_s_in_values(self, lang):
        """Убеждаемся что self.s() не просочилось в значения словаря."""
        bad = [k for k, v in core.STRINGS[lang].items()
               if isinstance(v, str) and "self.s(" in v]
        assert not bad, f"Язык {lang!r}: self.s() в значениях: {bad}"

    @pytest.mark.parametrize("lang", LANGUAGES)
    def test_all_values_are_strings(self, lang):
        bad = [k for k, v in core.STRINGS[lang].items() if not isinstance(v, str)]
        assert not bad, f"Язык {lang!r}: нестроковые значения: {bad}"

    def test_done_status_has_placeholders(self):
        template = core.STRINGS["ru"]["done_status"]
        assert "{rows}" in template
        assert "{cols}" in template
        assert "{name}" in template

    def test_done_msg_has_path_placeholder(self):
        for lang in self.LANGUAGES:
            msg = core.STRINGS[lang]["done_msg"]
            assert "{path}" in msg, f"Язык {lang!r}: done_msg без {{path}}"


# ═════════════════════════════════════════════════════════════════════════════
# Интеграционный тест: полный цикл CSV → XLSX → CSV
# ═════════════════════════════════════════════════════════════════════════════

class TestRoundTrip:

    def test_csv_xlsx_csv_round_trip(self, tmp):
        """Данные должны сохраняться при прямой + обратной конвертации."""
        original = tmp / "original.csv"
        original.write_text(
            "Ключ проблемы;Тема;Статус\n"
            "TASK-1;Первая задача;Открыта\n"
            "TASK-2;Вторая задача;Закрыта\n",
            encoding="utf-8-sig",
        )
        xlsx = str(tmp / "middle.xlsx")
        csv_out = str(tmp / "result.csv")

        core.convert([str(original)], xlsx)
        core.convert_xlsx_to_csv([xlsx], csv_out)

        df = pd.read_csv(csv_out, sep=";", encoding="utf-8")
        assert "TASK-1" in df.iloc[:, 0].values
        assert "TASK-2" in df.iloc[:, 0].values
        assert len(df) == 2
